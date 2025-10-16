# 影刀RPA - Seefar店铺信息抓取测试
# 功能：读取好店文件，遍历店铺ID，抓取店铺销售数据

import xbot
from xbot import excel
# 删除错误的导入，使用正确的xbot导入
import xbot.app.dialog as dialog
import xbot.selector as selector
from xbot import sleep

class SeefarCrawler:
    def __init__(self):
        self.excel_app = None
        self.browser = None
        self.results = []
        
        # 配置信息
        self.SEEFAR_URL_TEMPLATE = "https://seerfar.cn/admin/store-detail.html?storeId={shop_id}&platform=OZON"
        
        xbot.print("Seefar店铺信息抓取工具启动")


    #没有用的方法
    def select_excel_file(self, args=None):
        """获取Excel文件路径"""
        try:
            # 从流程参数中获取文件路径
            if args:
                # 优先获取对话框结果（推荐方式）
                dialog_result = args.get("dialog_result")
                if dialog_result:
                    xbot.print(f"✅ 从对话框结果获取到文件路径：{dialog_result}")
                    return dialog_result

                # 兼容其他参数名
                possible_keys = ["good_shop_file", "excel_file_path", "file_path", "selected_file_path"]
                for key in possible_keys:
                    file_path = args.get(key)
                    if file_path:
                        xbot.print(f"✅ 从参数'{key}'获取到文件路径：{file_path}")
                        return file_path

            # 如果没有获取到文件路径，提供详细的设置指导
            xbot.print("❌ 未从UI输出中获取到文件路径")
            xbot.print("💡 请在影刀RPA流程中按以下步骤设置：")
            xbot.print("   1. 添加'打开选择文件对话框'组件")
            xbot.print("   2. 设置输出变量名")
            xbot.print("   3. 在'调用模块'中设置参数：")
            xbot.print("      - 参数名：dialog_result")
            xbot.print("      - 参数值：选择对话框的输出变量")
            return None

        except Exception as e:
            xbot.print(f"获取文件路径失败：{str(e)}")
            return None
    
    def read_shop_ids(self, file_path):
        """读取Excel文件中的店铺ID"""
        try:
            # 使用openpyxl库读取Excel文件（跨平台兼容）
            from openpyxl import load_workbook

            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active  # 获取活动工作表

            shop_ids = []

            # 遍历第一列的所有行
            for row in worksheet.iter_rows(min_col=1, max_col=1, values_only=True):
                cell_value = row[0]

                if not cell_value or str(cell_value).strip() == "":
                    continue

                shop_id = str(cell_value).strip()
                if shop_id and shop_id != "店铺ID":
                    shop_ids.append(shop_id)

                # 限制最大读取行数，避免无限循环
                if len(shop_ids) >= 1000:
                    break

            workbook.close()
            xbot.print(f"✅ 使用openpyxl成功读取到 {len(shop_ids)} 个店铺ID")
            return shop_ids

        except ImportError:
            xbot.print("❌ openpyxl库未安装，尝试使用影刀内置Excel操作")
            return self._read_shop_ids_fallback(file_path)
        except Exception as e:
            xbot.print(f"❌ 使用openpyxl读取Excel文件失败：{str(e)}")
            xbot.print("🔄 尝试使用备用方法读取...")
            return self._read_shop_ids_fallback(file_path)

    def _read_shop_ids_fallback(self, file_path):
        """备用的Excel读取方法"""
        try:
            # 尝试使用CSV方式读取（如果文件是CSV格式）
            if file_path.lower().endswith('.csv'):
                import csv
                shop_ids = []
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if row and row[0].strip() and row[0].strip() != "店铺ID":
                            shop_ids.append(row[0].strip())
                xbot.print(f"✅ 使用CSV方式读取到 {len(shop_ids)} 个店铺ID")
                return shop_ids

            # 尝试使用pandas读取Excel
            try:
                import pandas as pd
                df = pd.read_excel(file_path, engine='openpyxl')
                shop_ids = []
                for value in df.iloc[:, 0]:  # 第一列
                    if pd.notna(value) and str(value).strip() != "店铺ID":
                        shop_ids.append(str(value).strip())
                xbot.print(f"✅ 使用pandas成功读取到 {len(shop_ids)} 个店铺ID")
                return shop_ids
            except ImportError:
                pass

            xbot.print("❌ 所有备用方法都失败了")
            xbot.print("💡 建议：")
            xbot.print("   1. 确保文件格式正确（Excel .xlsx 或 CSV .csv）")
            xbot.print("   2. 检查文件路径是否正确")
            xbot.print("   3. 尝试将Excel文件另存为CSV格式")
            return []

        except Exception as e:
            xbot.print(f"❌ 备用读取方法也失败了：{str(e)}")
            return []
    
    def init_browser(self):
        """初始化浏览器"""
        try:
            # ✅ 正确的影刀RPA浏览器创建API
            self.browser = xbot.web.create("https://seerfar.cn", "chrome", load_timeout=20)
            xbot.print("✅ 浏览器初始化成功")
            return True
        except Exception as e:
            xbot.print(f"❌ 浏览器初始化失败：{str(e)}")
            return False
    
    def crawl_shop_data(self, shop_id):
        """抓取单个店铺的数据"""
        try:
            url = self.SEEFAR_URL_TEMPLATE.format(shop_id=shop_id)
            self.browser.get(url)
            sleep(3)

            shop_data = {
                "shop_id": shop_id,
                "sales_amount": "未获取",
                "sales_volume": "未获取",
                "daily_avg_sales": "未获取"
            }

            try:
                sales_amount_element = selector.find_element(
                    self.browser,
                    "//span[contains(text(),'销售额')]/following-sibling::span"
                )
                if sales_amount_element:
                    shop_data["sales_amount"] = sales_amount_element.text
            except:
                pass

            try:
                sales_volume_element = selector.find_element(
                    self.browser,
                    "//span[contains(text(),'销量')]/following-sibling::span"
                )
                if sales_volume_element:
                    shop_data["sales_volume"] = sales_volume_element.text
            except:
                pass

            try:
                daily_avg_element = selector.find_element(
                    self.browser,
                    "//span[contains(text(),'日均')]/following-sibling::span"
                )
                if daily_avg_element:
                    shop_data["daily_avg_sales"] = daily_avg_element.text
            except:
                pass

            return shop_data

        except Exception as e:
            xbot.print(f"抓取店铺 {shop_id} 失败：{str(e)}")
            return {
                "shop_id": shop_id,
                "sales_amount": "抓取失败",
                "sales_volume": "抓取失败",
                "daily_avg_sales": "抓取失败"
            }
    
    def close_browser(self):
        """关闭浏览器"""
        if self.browser:
            try:
                self.browser.quit()
            except:
                pass
    
    def close_excel(self):
        """关闭Excel应用"""
        pass
    
    def output_results(self):
        """输出抓取结果"""
        if not self.results:
            xbot.print("没有抓取到任何数据")
            return

        for i, data in enumerate(self.results, 1):
            xbot.print(f"店铺 {i}: {data['shop_id']}")
            xbot.print(f"  销售额: {data['sales_amount']}")
            xbot.print(f"  销量: {data['sales_volume']}")
            xbot.print(f"  日均销量: {data['daily_avg_sales']}")

        xbot.print(f"总计处理了 {len(self.results)} 个店铺")

        # 通过args参数设置输出变量（如果提供了args参数）
        if hasattr(self, 'args') and self.args:
            try:
                self.args["crawl_count"] = len(self.results)
                self.args["crawl_results"] = str(self.results)
                xbot.print("已将结果设置到输出参数中")
            except Exception as e:
                xbot.print(f"设置输出参数失败：{str(e)}")
        else:
            xbot.print("未提供args参数，跳过结果输出设置")
    
    def run(self, args=None):
        """主运行流程"""
        try:
            # 保存args参数供其他方法使用
            self.args = args

            file_path = self.select_excel_file(args)
            if not file_path:
                return

            shop_ids = self.read_shop_ids(file_path)
            if not shop_ids:
                xbot.print("未读取到店铺ID")
                return

            if not self.init_browser():
                return

            for i, shop_id in enumerate(shop_ids, 1):
                xbot.print(f"处理第 {i}/{len(shop_ids)} 个店铺")

                shop_data = self.crawl_shop_data(shop_id)
                self.results.append(shop_data)

                try:
                    self.browser.close_current_tab()
                except:
                    pass

                sleep(2)

                if i >= 5:
                    break

            self.output_results()

        except Exception as e:
            xbot.print(f"程序执行异常：{str(e)}")

        finally:
            self.close_browser()
            self.close_excel()



# 模块定义完成，可被其他脚本导入使用