"""
Excel Compiler Module

This module compiles Excel workbooks into Python code for fast execution.
It extracts formulas, dependencies, and lookup tables from Excel files
and generates optimized Python code.
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, List, Tuple, Set, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import ast
import re


class ExcelCompiler:
    """Compiles Excel workbooks into Python code"""
    
    logger = logging.getLogger(__name__)
    
    def __init__(self):
        """Initialize Excel compiler"""
        self.workbook = None
        self.formulas: Dict[str, Dict[str, Any]] = {}  # sheet -> cell -> formula info
        self.dependencies: Dict[str, Set[str]] = {}  # cell -> dependencies
        self.constants: Dict[str, Any] = {}  # cell -> constant value
        self.lookup_tables: Dict[str, List[List[Any]]] = {}  # sheet -> table data
        
    def compile_workbook(self, excel_path: str) -> Dict[str, Any]:
        """
        Compile an Excel workbook into Python structures
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            Dictionary containing compiled data
        """
        self.logger.info(f"Compiling Excel workbook: {excel_path}")
        
        # Load workbook
        self.workbook = openpyxl.load_workbook(excel_path, data_only=False)
        
        # Extract all data
        self._extract_formulas()
        self._extract_constants()
        self._extract_lookup_tables()
        self._analyze_dependencies()
        
        # Generate compilation result
        result = {
            "formulas": self.formulas,
            "dependencies": {k: list(v) for k, v in self.dependencies.items()},
            "constants": self.constants,
            "lookup_tables": self.lookup_tables,
            "metadata": {
                "source_file": excel_path,
                "sheets": list(self.workbook.sheetnames),
                "total_formulas": sum(len(sheet_formulas) for sheet_formulas in self.formulas.values()),
                "total_constants": len(self.constants)
            }
        }
        
        self.logger.info(f"Compilation complete: {result['metadata']}")
        return result
        
    def generate_python_code(self, compiled_data: Dict[str, Any], output_path: str) -> None:
        """
        Generate Python code from compiled data
        
        Args:
            compiled_data: Compiled Excel data
            output_path: Path for output Python file
        """
        self.logger.info(f"Generating Python code to: {output_path}")
        
        lines = []
        lines.append('"""')
        lines.append('Auto-generated Excel calculation rules')
        lines.append(f'Source: {compiled_data["metadata"]["source_file"]}')
        lines.append('"""')
        lines.append('')
        lines.append('from typing import Dict, Any, Optional')
        lines.append('import math')
        lines.append('')
        
        # Generate constants
        lines.append('# Constants from Excel')
        lines.append('EXCEL_CONSTANTS = {')
        for cell, value in compiled_data["constants"].items():
            if isinstance(value, str):
                lines.append(f'    "{cell}": "{value}",')
            else:
                lines.append(f'    "{cell}": {value},')
        lines.append('}')
        lines.append('')
        
        # Generate lookup tables
        lines.append('# Lookup tables from Excel')
        lines.append('LOOKUP_TABLES = {')
        for sheet, table in compiled_data["lookup_tables"].items():
            lines.append(f'    "{sheet}": [')
            for row in table:
                lines.append(f'        {row},')
            lines.append('    ],')
        lines.append('}')
        lines.append('')
        
        # Generate formula functions
        lines.append('# Excel formulas as Python functions')
        lines.append('class ExcelCalculations:')
        lines.append('    """Excel formula calculations"""')
        lines.append('    ')
        lines.append('    def __init__(self):')
        lines.append('        self.cell_cache = {}')
        lines.append('        self.constants = EXCEL_CONSTANTS.copy()')
        lines.append('        self.lookup_tables = LOOKUP_TABLES')
        lines.append('    ')
        
        # Generate calculation methods
        for sheet, formulas in compiled_data["formulas"].items():
            for cell, formula_info in formulas.items():
                method_name = self._cell_to_method_name(sheet, cell)
                lines.append(f'    def {method_name}(self, inputs: Dict[str, Any]) -> Any:')
                lines.append(f'        """Calculate {sheet}!{cell}"""')
                
                # Convert Excel formula to Python
                python_code = self._formula_to_python(formula_info["formula"])
                lines.append(f'        return {python_code}')
                lines.append('    ')
        
        # Generate main calculation method
        lines.append('    def calculate_profit(self, inputs: Dict[str, Any]) -> Dict[str, Any]:')
        lines.append('        """Main profit calculation entry point"""')
        lines.append('        # Reset cache for new calculation')
        lines.append('        self.cell_cache = {}')
        lines.append('        ')
        lines.append('        # Perform calculations')
        lines.append('        result = {}')
        lines.append('        ')
        lines.append('        # Calculate profit (example - adjust based on Excel structure)')
        lines.append('        # This should call the appropriate cell calculation methods')
        lines.append('        ')
        lines.append('        return result')
        
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
            
        self.logger.info("Python code generation complete")
        
    def _extract_formulas(self):
        """Extract all formulas from the workbook"""
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            self.formulas[sheet_name] = {}
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).startswith('='):
                        self.formulas[sheet_name][cell.coordinate] = {
                            "formula": cell.value[1:],  # Remove '=' prefix
                            "address": f"{sheet_name}!{cell.coordinate}"
                        }
                        
    def _extract_constants(self):
        """Extract all constant values (non-formulas)"""
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and not str(cell.value).startswith('='):
                        address = f"{sheet_name}!{cell.coordinate}"
                        self.constants[address] = cell.value
                        
    def _extract_lookup_tables(self):
        """Extract lookup tables (shipping rates, etc.)"""
        # Extract specific known tables
        if "UNI运费" in self.workbook.sheetnames:
            self._extract_shipping_table()
            
    def _extract_shipping_table(self):
        """Extract shipping rate table"""
        sheet = self.workbook["UNI运费"]
        table_data = []
        
        # Skip header and read data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Check if row has data
                table_data.append(list(row))
                
        self.lookup_tables["shipping_rates"] = table_data
        
    def _analyze_dependencies(self):
        """Analyze formula dependencies"""
        for sheet_name, formulas in self.formulas.items():
            for cell, formula_info in formulas.items():
                address = formula_info["address"]
                self.dependencies[address] = self._extract_dependencies(formula_info["formula"])
                
    def _extract_dependencies(self, formula: str) -> Set[str]:
        """Extract cell references from a formula"""
        # Simple regex for cell references (can be improved)
        pattern = r'([A-Z]+\d+)'
        matches = re.findall(pattern, formula)
        return set(matches)
        
    def _cell_to_method_name(self, sheet: str, cell: str) -> str:
        """Convert sheet and cell to Python method name"""
        sheet_clean = re.sub(r'[^a-zA-Z0-9]', '_', sheet).lower()
        cell_clean = cell.lower()
        return f"calc_{sheet_clean}_{cell_clean}"
        
    def _formula_to_python(self, formula: str) -> str:
        """Convert Excel formula to Python expression"""
        # This is a simplified conversion - real implementation would be more complex
        python_expr = formula
        
        # Replace Excel functions with Python equivalents
        replacements = {
            'SUM': 'sum',
            'AVERAGE': 'lambda x: sum(x)/len(x)',
            'IF': 'lambda c,t,f: t if c else f',
            'MAX': 'max',
            'MIN': 'min',
            'ROUND': 'round',
        }
        
        for excel_func, python_func in replacements.items():
            python_expr = python_expr.replace(excel_func, python_func)
            
        # Replace cell references with lookups
        # This is simplified - real implementation needs proper parsing
        pattern = r'([A-Z]+\d+)'
        python_expr = re.sub(pattern, r'self.get_cell_value("\1", inputs)', python_expr)
        
        return python_expr


class ExcelCompilerCLI:
    """CLI interface for Excel compilation"""
    
    @staticmethod
    def compile_excel(excel_path: str, output_dir: str = None) -> str:
        """
        Compile Excel file to Python code
        
        Args:
            excel_path: Path to Excel file
            output_dir: Output directory (defaults to same as Excel)
            
        Returns:
            Path to generated Python file
        """
        compiler = ExcelCompiler()
        
        # Compile workbook
        compiled_data = compiler.compile_workbook(excel_path)
        
        # Determine output paths
        excel_path = Path(excel_path)
        if output_dir is None:
            output_dir = excel_path.parent
        else:
            output_dir = Path(output_dir)
            
        # Save compiled data
        json_path = output_dir / f"{excel_path.stem}_compiled.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(compiled_data, f, indent=2, default=str)
            
        # Generate Python code
        python_path = output_dir / f"{excel_path.stem}_rules.py"
        compiler.generate_python_code(compiled_data, str(python_path))
        
        return str(python_path)


if __name__ == "__main__":
    # CLI usage
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python excel_compiler.py <excel_file> [output_dir]")
        sys.exit(1)
        
    excel_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        output_file = ExcelCompilerCLI.compile_excel(excel_file, output_dir)
        print(f"Compilation successful: {output_file}")
    except Exception as e:
        print(f"Compilation failed: {e}")
        sys.exit(1)