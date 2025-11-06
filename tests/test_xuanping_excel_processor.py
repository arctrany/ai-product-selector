"""
测试好店筛选系统的Excel处理模块

测试Excel文件读写、店铺数据处理和利润计算功能。
"""

import pytest
import tempfile
import os
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook

from apps.xuanping.common.excel_processor import ExcelStoreProcessor, ExcelProfitProcessor
from apps.xuanping.common.models import (
    StoreInfo, ProductInfo, PriceCalculationResult, StoreAnalysisResult,
    StoreStatus, GoodStoreFlag
)
from apps.xuanping.common.config import GoodStoreSelectorConfig


class TestExcelStoreProcessor:
    """测试Excel店铺处理器"""
    
    def setup_method(self):
        """测试前的设置"""
        self.config = GoodStoreSelectorConfig()
        self.processor = ExcelStoreProcessor(self.config)
    
    def create_test_excel_file(self, store_data=None):
        """创建测试用的Excel文件"""
        wb = Workbook()
        ws = wb.active
        
        # 设置表头
        headers = ['店铺ID', '是否为好店', '状态', '销售额', '订单数', '评估时间']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 添加测试数据
        if store_data:
            for row, data in enumerate(store_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
        
        # 保存到临时文件
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        return temp_file.name
    
    def test_excel_store_processor_initialization(self):
        """测试Excel店铺处理器初始化"""
        assert self.processor.config is not None
        assert self.processor.logger is not None
        
        # 测试使用默认配置初始化
        processor_default = ExcelStoreProcessor()
        assert processor_default.config is not None
    
    def test_read_store_list_success(self):
        """测试成功读取店铺列表"""
        # 创建测试数据
        store_data = [
            ['STORE001', '', '待处理', 600000, 300, ''],
            ['STORE002', '是', '已完成', 800000, 400, '2024-01-01'],
            ['STORE003', '否', '已完成', 200000, 100, '2024-01-02']
        ]
        
        excel_file = self.create_test_excel_file(store_data)
        
        try:
            stores = self.processor.read_store_list(excel_file)
            
            assert len(stores) == 3
            
            # 验证第一个店铺
            store1 = stores[0]
            assert store1.store_id == 'STORE001'
            assert store1.good_store_flag == GoodStoreFlag.UNKNOWN
            assert store1.status == StoreStatus.PENDING
            
            # 验证第二个店铺
            store2 = stores[1]
            assert store2.store_id == 'STORE002'
            assert store2.good_store_flag == GoodStoreFlag.YES
            assert store2.status == StoreStatus.COMPLETED
            
            # 验证第三个店铺
            store3 = stores[2]
            assert store3.store_id == 'STORE003'
            assert store3.good_store_flag == GoodStoreFlag.NO
            assert store3.status == StoreStatus.COMPLETED
            
        finally:
            os.unlink(excel_file)
    
    def test_read_store_list_file_not_found(self):
        """测试读取不存在的文件"""
        with pytest.raises(FileNotFoundError):
            self.processor.read_store_list("nonexistent_file.xlsx")
    
    def test_read_store_list_empty_file(self):
        """测试读取空文件"""
        excel_file = self.create_test_excel_file([])
        
        try:
            stores = self.processor.read_store_list(excel_file)
            assert len(stores) == 0
        finally:
            os.unlink(excel_file)
    
    def test_update_store_status(self):
        """测试更新店铺状态"""
        store_data = [
            ['STORE001', '', '待处理', 600000, 300, ''],
            ['STORE002', '', '待处理', 800000, 400, '']
        ]
        
        excel_file = self.create_test_excel_file(store_data)
        
        try:
            # 更新店铺状态
            self.processor.update_store_status(excel_file, 'STORE001', StoreStatus.PROCESSING)
            
            # 重新读取验证
            stores = self.processor.read_store_list(excel_file)
            store1 = next(s for s in stores if s.store_id == 'STORE001')
            assert store1.status == StoreStatus.PROCESSING
            
        finally:
            os.unlink(excel_file)
    
    def test_update_store_evaluation_result(self):
        """测试更新店铺评估结果"""
        store_data = [
            ['STORE001', '', '待处理', 600000, 300, '']
        ]
        
        excel_file = self.create_test_excel_file(store_data)
        
        try:
            # 创建评估结果 - 使用字典格式模拟评估结果
            evaluation_data = {
                'store_id': 'STORE001',
                'is_good_store': True,
                'total_products_checked': 50,
                'profitable_products_count': 15,
                'profit_rate': 30.0,
                'evaluation_time': '2024-01-01 12:00:00'
            }

            # 更新评估结果
            self.processor.update_store_status(excel_file, 'STORE001', StoreStatus.COMPLETED)
            # 模拟更新好店标记
            self.processor.batch_update_stores(excel_file, [('STORE001', StoreStatus.COMPLETED, GoodStoreFlag.YES)])
            
            # 重新读取验证
            stores = self.processor.read_store_list(excel_file)
            store1 = stores[0]
            assert store1.good_store_flag == GoodStoreFlag.YES
            assert store1.status == StoreStatus.COMPLETED
            
        finally:
            os.unlink(excel_file)
    
    def test_get_pending_stores(self):
        """测试获取待处理店铺"""
        store_data = [
            ['STORE001', '', '待处理', 600000, 300, ''],
            ['STORE002', '是', '已完成', 800000, 400, '2024-01-01'],
            ['STORE003', '', '处理中', 700000, 350, '']
        ]
        
        excel_file = self.create_test_excel_file(store_data)
        
        try:
            pending_stores = self.processor.get_pending_stores(excel_file)
            
            assert len(pending_stores) == 1
            assert pending_stores[0].store_id == 'STORE001'
            assert pending_stores[0].status == StoreStatus.PENDING
            
        finally:
            os.unlink(excel_file)
    
    def test_batch_update_stores(self):
        """测试批量更新店铺"""
        store_data = [
            ['STORE001', '', '待处理', 600000, 300, ''],
            ['STORE002', '', '待处理', 800000, 400, '']
        ]
        
        excel_file = self.create_test_excel_file(store_data)
        
        try:
            # 创建批量更新数据
            updates = [
                ('STORE001', StoreStatus.COMPLETED, GoodStoreFlag.YES),
                ('STORE002', StoreStatus.COMPLETED, GoodStoreFlag.NO)
            ]
            
            # 执行批量更新
            self.processor.batch_update_stores(excel_file, updates)
            
            # 验证更新结果
            stores = self.processor.read_store_list(excel_file)
            
            store1 = next(s for s in stores if s.store_id == 'STORE001')
            assert store1.status == StoreStatus.COMPLETED
            assert store1.good_store_flag == GoodStoreFlag.YES
            
            store2 = next(s for s in stores if s.store_id == 'STORE002')
            assert store2.status == StoreStatus.COMPLETED
            assert store2.good_store_flag == GoodStoreFlag.NO
            
        finally:
            os.unlink(excel_file)
    
    def test_error_handling(self):
        """测试错误处理"""
        # 测试无效的Excel文件
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as f:
            f.write(b"invalid excel content")
            temp_file = f.name
        
        try:
            with pytest.raises(Exception):
                self.processor.read_store_list(temp_file)
        finally:
            os.unlink(temp_file)


class TestExcelProfitProcessor:
    """测试Excel利润处理器"""
    
    def setup_method(self):
        """测试前的设置"""
        self.config = GoodStoreSelectorConfig()
        self.processor = ExcelProfitProcessor(self.config)
    
    @patch('apps.xuanping.common.excel_processor.ExcelCalculator')
    def test_excel_profit_processor_initialization(self, mock_calculator):
        """测试Excel利润处理器初始化"""
        assert self.processor.config is not None
        assert self.processor.logger is not None
        mock_calculator.assert_called_once()
    
    @patch('apps.xuanping.common.excel_processor.ExcelCalculator')
    def test_calculate_product_profit(self, mock_calculator):
        """测试计算商品利润"""
        # 模拟Excel计算器
        mock_calc_instance = mock_calculator.return_value
        mock_calc_instance.calculate_profit.return_value = {
            'profit_amount': 15.5,
            'profit_rate': 25.8,
            'is_profitable': True
        }
        
        # 创建测试商品
        product = ProductInfo(
            product_id='PROD001',
            title='测试商品',
            price_rub=2000.0,
            image_url='http://example.com/image.jpg'
        )
        
        # 创建价格计算结果
        price_result = PriceCalculationResult(
            real_selling_price=150.0,
            product_pricing=142.5,
            profit_amount=7.5,
            profit_rate=5.0,
            is_profitable=False,
            calculation_details={}
        )
        
        # 计算利润
        result = self.processor.calculate_product_profit(product, price_result, source_price=120.0)
        
        assert result['profit_amount'] == 15.5
        assert result['profit_rate'] == 25.8
        assert result['is_profitable'] == True
        
        # 验证调用参数
        mock_calc_instance.calculate_profit.assert_called_once()
        call_args = mock_calc_instance.calculate_profit.call_args[1]
        assert call_args['selling_price'] == 142.5
        assert call_args['source_price'] == 120.0
    
    @patch('apps.xuanping.common.excel_processor.ExcelCalculator')
    def test_batch_calculate_profits(self, mock_calculator):
        """测试批量计算利润"""
        # 模拟Excel计算器
        mock_calc_instance = mock_calculator.return_value
        mock_calc_instance.calculate_profit.side_effect = [
            {'profit_amount': 15.5, 'profit_rate': 25.8, 'is_profitable': True},
            {'profit_amount': -5.0, 'profit_rate': -8.3, 'is_profitable': False}
        ]
        
        # 创建测试数据
        products_data = [
            {
                'product': ProductInfo(
                    product_id='PROD001',
                    title='商品1',
                    price_rub=2000.0,
                    image_url='http://example.com/1.jpg'
                ),
                'price_result': PriceCalculationResult(
                    real_selling_price=150.0,
                    product_pricing=142.5,
                    profit_amount=7.5,
                    profit_rate=5.0,
                    is_profitable=False,
                    calculation_details={}
                ),
                'source_price': 120.0
            },
            {
                'product': ProductInfo(
                    product_id='PROD002',
                    title='商品2',
                    price_rub=1500.0,
                    image_url='http://example.com/2.jpg'
                ),
                'price_result': PriceCalculationResult(
                    real_selling_price=112.5,
                    product_pricing=106.9,
                    profit_amount=6.9,
                    profit_rate=6.1,
                    is_profitable=False,
                    calculation_details={}
                ),
                'source_price': 110.0
            }
        ]
        
        # 批量计算
        results = self.processor.batch_calculate_profits(products_data)
        
        assert len(results) == 2
        assert results[0]['profit_amount'] == 15.5
        assert results[0]['is_profitable'] == True
        assert results[1]['profit_amount'] == -5.0
        assert results[1]['is_profitable'] == False
        
        # 验证调用次数
        assert mock_calc_instance.calculate_profit.call_count == 2
    
    @patch('apps.xuanping.common.excel_processor.ExcelCalculator')
    def test_get_profit_statistics(self, mock_calculator):
        """测试获取利润统计"""
        # 创建测试利润结果
        profit_results = [
            {'profit_amount': 15.5, 'profit_rate': 25.8, 'is_profitable': True},
            {'profit_amount': 8.2, 'profit_rate': 12.3, 'is_profitable': True},
            {'profit_amount': -5.0, 'profit_rate': -8.3, 'is_profitable': False},
            {'profit_amount': 22.1, 'profit_rate': 35.7, 'is_profitable': True}
        ]
        
        stats = self.processor.get_profit_statistics(profit_results)
        
        assert stats['total_products'] == 4
        assert stats['profitable_products'] == 3
        assert stats['unprofitable_products'] == 1
        assert stats['profit_rate'] == 75.0  # 3/4 * 100
        assert stats['average_profit_amount'] == 10.2  # (15.5 + 8.2 - 5.0 + 22.1) / 4
        assert stats['average_profit_rate'] == 16.375  # (25.8 + 12.3 - 8.3 + 35.7) / 4
        assert stats['max_profit_amount'] == 22.1
        assert stats['min_profit_amount'] == -5.0
    
    @patch('apps.xuanping.common.excel_processor.ExcelCalculator')
    def test_filter_profitable_products(self, mock_calculator):
        """测试筛选有利润商品"""
        # 创建测试数据
        products_with_profits = [
            {
                'product': ProductInfo(product_id='PROD001', title='商品1', price_rub=2000.0, image_url=''),
                'profit_result': {'profit_amount': 15.5, 'profit_rate': 25.8, 'is_profitable': True}
            },
            {
                'product': ProductInfo(product_id='PROD002', title='商品2', price_rub=1500.0, image_url=''),
                'profit_result': {'profit_amount': -5.0, 'profit_rate': -8.3, 'is_profitable': False}
            },
            {
                'product': ProductInfo(product_id='PROD003', title='商品3', price_rub=1800.0, image_url=''),
                'profit_result': {'profit_amount': 8.2, 'profit_rate': 12.3, 'is_profitable': True}
            }
        ]
        
        profitable = self.processor.filter_profitable_products(products_with_profits)
        
        assert len(profitable) == 2
        assert profitable[0]['product'].product_id == 'PROD001'
        assert profitable[1]['product'].product_id == 'PROD003'
    
    @patch('apps.xuanping.common.excel_processor.ExcelCalculator')
    def test_error_handling(self, mock_calculator):
        """测试错误处理"""
        # 模拟Excel计算器抛出异常
        mock_calc_instance = mock_calculator.return_value
        mock_calc_instance.calculate_profit.side_effect = Exception("计算错误")
        
        product = ProductInfo(
            product_id='PROD001',
            title='测试商品',
            price_rub=2000.0,
            image_url='http://example.com/image.jpg'
        )
        
        price_result = PriceCalculationResult(
            real_selling_price=150.0,
            product_pricing=142.5,
            profit_amount=7.5,
            profit_rate=5.0,
            is_profitable=False,
            calculation_details={}
        )
        
        # 应该返回默认的无利润结果
        result = self.processor.calculate_product_profit(product, price_result, source_price=120.0)
        
        assert result['profit_amount'] == 0.0
        assert result['profit_rate'] == 0.0
        assert result['is_profitable'] == False


class TestExcelProcessorIntegration:
    """测试Excel处理器的集成场景"""
    
    def test_complete_store_processing_workflow(self):
        """测试完整的店铺处理工作流"""
        config = GoodStoreSelectorConfig()
        store_processor = ExcelStoreProcessor(config)
        
        # 创建测试Excel文件
        store_data = [
            ['STORE001', '', '待处理', 600000, 300, ''],
            ['STORE002', '', '待处理', 800000, 400, '']
        ]
        
        wb = Workbook()
        ws = wb.active
        
        # 设置表头
        headers = ['店铺ID', '是否为好店', '状态', '销售额', '订单数', '评估时间']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 添加数据
        for row, data in enumerate(store_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 保存文件
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        try:
            # 1. 读取待处理店铺
            pending_stores = store_processor.get_pending_stores(temp_file.name)
            assert len(pending_stores) == 2
            
            # 2. 更新第一个店铺为处理中
            store_processor.update_store_status(temp_file.name, 'STORE001', StoreStatus.PROCESSING)
            
            # 3. 完成第一个店铺的评估
            store_processor.batch_update_stores(temp_file.name, [('STORE001', StoreStatus.COMPLETED, GoodStoreFlag.YES)])
            
            # 4. 验证最终状态
            final_stores = store_processor.read_store_list(temp_file.name)
            
            store1 = next(s for s in final_stores if s.store_id == 'STORE001')
            assert store1.status == StoreStatus.COMPLETED
            assert store1.good_store_flag == GoodStoreFlag.YES
            
            store2 = next(s for s in final_stores if s.store_id == 'STORE002')
            assert store2.status == StoreStatus.PENDING
            assert store2.good_store_flag == GoodStoreFlag.UNKNOWN
            
        finally:
            os.unlink(temp_file.name)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])