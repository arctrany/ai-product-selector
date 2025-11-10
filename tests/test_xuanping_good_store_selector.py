"""
æµ‹è¯•å¥½åº—ç­›é€‰ç³»ç»Ÿçš„ä¸»æµç¨‹ç¼–æ’

æµ‹è¯•å®Œæ•´çš„å¥½åº—ç­›é€‰å·¥ä½œæµç¨‹å’Œç³»ç»Ÿé›†æˆã€‚
"""

import pytest
import tempfile
import os
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook

from apps.xuanping.common.good_store_selector import GoodStoreSelector
from apps.xuanping.common.models import (
    StoreInfo, ProductInfo, PriceCalculationResult, ProductAnalysisResult,
    StoreAnalysisResult, BatchProcessingResult, StoreStatus, GoodStoreFlag
)
from apps.xuanping.common.config import GoodStoreSelectorConfig


class TestGoodStoreSelector:
    """æµ‹è¯•å¥½åº—ç­›é€‰ç³»ç»Ÿä¸»æ§åˆ¶å™¨"""
    
    def setup_method(self):
        """æµ‹è¯•å‰çš„è®¾ç½®"""
        self.config = GoodStoreSelectorConfig()
        self.selector = GoodStoreSelector(self.config)
    
    def test_good_store_selector_initialization(self):
        """æµ‹è¯•å¥½åº—ç­›é€‰ç³»ç»Ÿåˆå§‹åŒ–"""
        assert self.selector.config is not None
        assert self.selector.logger is not None
        assert self.selector.excel_processor is not None
        assert self.selector.pricing_calculator is not None
        assert self.selector.profit_evaluator is not None
        assert self.selector.store_evaluator is not None
        assert self.selector.source_matcher is not None
        
        # æµ‹è¯•ä½¿ç”¨é»˜è®¤é…ç½®åˆå§‹åŒ–
        selector_default = GoodStoreSelector()
        assert selector_default.config is not None
    
    def create_test_excel_file(self, store_data=None):
        """åˆ›å»ºæµ‹è¯•ç”¨çš„Excelæ–‡ä»¶"""
        wb = Workbook()
        ws = wb.active
        
        # è®¾ç½®è¡¨å¤´
        headers = ['åº—é“ºID', 'æ˜¯å¦ä¸ºå¥½åº—', 'çŠ¶æ€', 'é”€å”®é¢', 'è®¢å•æ•°', 'è¯„ä¼°æ—¶é—´']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # æ·»åŠ æµ‹è¯•æ•°æ®
        if store_data:
            for row, data in enumerate(store_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
        
        # ä¿å­˜åˆ°ä¸´æ—¶æ–‡ä»¶
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        return temp_file.name
    
    @patch('apps.xuanping.common.good_store_selector.SeerfarScraper')
    @patch('apps.xuanping.common.good_store_selector.OzonScraper')
    @patch('apps.xuanping.common.good_store_selector.ErpPluginScraper')
    def test_process_single_store_success(self, mock_erp_scraper, mock_ozon_scraper, mock_seerfar_scraper):
        """æµ‹è¯•æˆåŠŸå¤„ç†å•ä¸ªåº—é“º"""
        # æ¨¡æ‹ŸæŠ“å–å™¨
        mock_seerfar = mock_seerfar_scraper.return_value
        mock_ozon = mock_ozon_scraper.return_value
        mock_erp = mock_erp_plugin_scraper.return_value
        
        # æ¨¡æ‹Ÿåº—é“ºé”€å”®æ•°æ®æŠ“å–
        mock_seerfar.scrape_store_sales_data.return_value = {
            'sales_30days': 600000.0,
            'orders_30days': 300,
            'success': True
        }
        
        # æ¨¡æ‹Ÿå•†å“åˆ—è¡¨æŠ“å–
        mock_seerfar.scrape_store_products.return_value = [
            ProductInfo(
                product_id='PROD001',
                title='æµ‹è¯•å•†å“1',
                price_rub=2000.0,
                image_url='http://example.com/1.jpg'
            ),
            ProductInfo(
                product_id='PROD002',
                title='æµ‹è¯•å•†å“2',
                price_rub=1500.0,
                image_url='http://example.com/2.jpg'
            )
        ]
        
        # æ¨¡æ‹ŸOZONä»·æ ¼æŠ“å–
        mock_ozon.scrape_product_prices.return_value = {
            'green_price_rub': 1800.0,
            'black_price_rub': 2000.0,
            'success': True
        }
        
        # æ¨¡æ‹ŸERPæ’ä»¶æ•°æ®æŠ“å–
        mock_erp.scrape_product_attributes.return_value = {
            'commission_rate': 12.0,
            'weight': 500.0,
            'success': True
        }
        
        # åˆ›å»ºæµ‹è¯•åº—é“º
        store = StoreInfo(
            store_id='TEST_STORE_001',
            status=StoreStatus.PENDING
        )
        
        # å¤„ç†åº—é“º
        result = self.selector.process_single_store(store)
        
        # éªŒè¯ç»“æœ
        assert isinstance(result, StoreAnalysisResult)
        assert result.store_info.store_id == 'TEST_STORE_001'
        assert len(result.products) > 0
        
        # éªŒè¯æŠ“å–å™¨è¢«è°ƒç”¨
        mock_seerfar.scrape_store_sales_data.assert_called_once()
        mock_seerfar.scrape_store_products.assert_called_once()
    
    @patch('apps.xuanping.common.good_store_selector.SeerfarScraper')
    def test_process_single_store_initial_filter_fail(self, mock_seerfar_scraper):
        """æµ‹è¯•åº—é“ºåˆç­›å¤±è´¥"""
        # æ¨¡æ‹ŸæŠ“å–å™¨
        mock_seerfar = mock_seerfar_scraper.return_value
        
        # æ¨¡æ‹Ÿé”€å”®æ•°æ®ä¸æ»¡è¶³åˆç­›æ¡ä»¶
        mock_seerfar.scrape_store_sales_data.return_value = {
            'sales_30days': 100000.0,  # ä½äºé˜ˆå€¼
            'orders_30days': 50,       # ä½äºé˜ˆå€¼
            'success': True
        }
        
        # åˆ›å»ºæµ‹è¯•åº—é“º
        store = StoreInfo(
            store_id='TEST_STORE_002',
            status=StoreStatus.PENDING
        )
        
        # å¤„ç†åº—é“º
        result = self.selector.process_single_store(store)
        
        # éªŒè¯ç»“æœ
        assert isinstance(result, StoreAnalysisResult)
        assert result.store_info.store_id == 'TEST_STORE_002'
        assert len(result.products) == 0  # æ²¡æœ‰å•†å“æ•°æ®
        assert result.total_products == 0
        assert result.is_good_store == False
    
    @patch('apps.xuanping.common.good_store_selector.SeerfarScraper')
    def test_process_single_store_scraping_error(self, mock_seerfar_scraper):
        """æµ‹è¯•æŠ“å–è¿‡ç¨‹ä¸­çš„é”™è¯¯å¤„ç†"""
        # æ¨¡æ‹ŸæŠ“å–å™¨
        mock_seerfar = mock_seerfar_scraper.return_value
        
        # æ¨¡æ‹ŸæŠ“å–å¤±è´¥
        mock_seerfar.scrape_store_sales_data.side_effect = Exception("ç½‘ç»œé”™è¯¯")
        
        # åˆ›å»ºæµ‹è¯•åº—é“º
        store = StoreInfo(
            store_id='TEST_STORE_003',
            status=StoreStatus.PENDING
        )
        
        # å¤„ç†åº—é“ºåº”è¯¥ä¸ä¼šæŠ›å‡ºå¼‚å¸¸
        result = self.selector.process_single_store(store)
        
        # éªŒè¯é”™è¯¯å¤„ç†
        assert isinstance(result, StoreAnalysisResult)
        assert result.store_info.store_id == 'TEST_STORE_003'
        assert result.has_error == True
    
    def test_process_stores_from_excel(self):
        """æµ‹è¯•ä»Excelæ–‡ä»¶å¤„ç†åº—é“º"""
        # åˆ›å»ºæµ‹è¯•Excelæ–‡ä»¶
        store_data = [
            ['STORE001', '', 'å¾…å¤„ç†', 600000, 300, ''],
            ['STORE002', '', 'å¾…å¤„ç†', 800000, 400, '']
        ]
        
        excel_file = self.create_test_excel_file(store_data)
        
        try:
            # æ¨¡æ‹Ÿprocess_single_storeæ–¹æ³•
            with patch.object(self.selector, 'process_single_store') as mock_process:
                mock_process.return_value = StoreAnalysisResult(
                    store_info=StoreInfo(store_id='STORE001'),
                    products=[],
                    total_products=0,
                    profitable_products=0,
                    profit_rate=0.0,
                    is_good_store=False
                )
                
                # å¤„ç†Excelæ–‡ä»¶
                result = self.selector.process_stores_from_excel(excel_file)
                
                # éªŒè¯ç»“æœ
                assert isinstance(result, BatchProcessingResult)
                assert result.total_stores == 2
                assert result.processed_stores == 2
                
                # éªŒè¯process_single_storeè¢«è°ƒç”¨äº†2æ¬¡
                assert mock_process.call_count == 2
                
        finally:
            os.unlink(excel_file)
    
    def test_process_stores_from_excel_with_limit(self):
        """æµ‹è¯•é™åˆ¶å¤„ç†åº—é“ºæ•°é‡"""
        # åˆ›å»ºæµ‹è¯•Excelæ–‡ä»¶
        store_data = [
            ['STORE001', '', 'å¾…å¤„ç†', 600000, 300, ''],
            ['STORE002', '', 'å¾…å¤„ç†', 800000, 400, ''],
            ['STORE003', '', 'å¾…å¤„ç†', 700000, 350, '']
        ]
        
        excel_file = self.create_test_excel_file(store_data)
        
        try:
            # æ¨¡æ‹Ÿprocess_single_storeæ–¹æ³•
            with patch.object(self.selector, 'process_single_store') as mock_process:
                mock_process.return_value = StoreAnalysisResult(
                    store_info=StoreInfo(store_id='STORE001'),
                    products=[],
                    total_products=0,
                    profitable_products=0,
                    profit_rate=0.0,
                    is_good_store=False
                )
                
                # é™åˆ¶å¤„ç†2ä¸ªåº—é“º
                result = self.selector.process_stores_from_excel(excel_file, max_stores=2)
                
                # éªŒè¯ç»“æœ
                assert isinstance(result, BatchProcessingResult)
                assert result.total_stores == 3  # æ€»å…±3ä¸ªåº—é“º
                assert result.processed_stores == 2  # åªå¤„ç†äº†2ä¸ª
                
                # éªŒè¯process_single_storeè¢«è°ƒç”¨äº†2æ¬¡
                assert mock_process.call_count == 2
                
        finally:
            os.unlink(excel_file)
    
    def test_validate_store_filter_conditions(self):
        """æµ‹è¯•åº—é“ºç­›é€‰æ¡ä»¶éªŒè¯"""
        # æµ‹è¯•æ»¡è¶³æ¡ä»¶çš„é”€å”®æ•°æ®
        sales_data_pass = {
            'sales_30days': 600000.0,
            'orders_30days': 300
        }
        
        result = self.selector._validate_store_filter_conditions(sales_data_pass)
        assert result == True
        
        # æµ‹è¯•ä¸æ»¡è¶³é”€å”®é¢æ¡ä»¶
        sales_data_fail_sales = {
            'sales_30days': 400000.0,  # ä½äº500000é˜ˆå€¼
            'orders_30days': 300
        }
        
        result = self.selector._validate_store_filter_conditions(sales_data_fail_sales)
        assert result == False
        
        # æµ‹è¯•ä¸æ»¡è¶³è®¢å•æ•°æ¡ä»¶
        sales_data_fail_orders = {
            'sales_30days': 600000.0,
            'orders_30days': 200  # ä½äº250é˜ˆå€¼
        }
        
        result = self.selector._validate_store_filter_conditions(sales_data_fail_orders)
        assert result == False
    
    # ğŸ”§ å·²åˆ é™¤ï¼š_convert_image_url_to_product_url æ–¹æ³•å·²åˆ é™¤ï¼Œå› ä¸ºå•†å“é¡µé¢é€šè¿‡ç‚¹å‡»æ‰“å¼€ï¼Œä¸éœ€è¦URLè½¬æ¢
    
    def test_generate_processing_summary(self):
        """æµ‹è¯•ç”Ÿæˆå¤„ç†æ‘˜è¦"""
        # åˆ›å»ºæµ‹è¯•ç»“æœ
        result = BatchProcessingResult(
            total_stores=10,
            processed_stores=8,
            good_stores=3,
            failed_stores=2,
            start_time='2024-01-01 10:00:00',
            end_time='2024-01-01 12:00:00',
            processing_time_seconds=7200.0
        )
        
        summary = self.selector.generate_processing_summary(result)
        
        # éªŒè¯æ‘˜è¦å†…å®¹
        assert "æ€»åº—é“ºæ•°: 10" in summary
        assert "å¤„ç†æˆåŠŸ: 8" in summary
        assert "å¥½åº—æ•°é‡: 3" in summary
        assert "å¤±è´¥æ•°é‡: 2" in summary
        assert "å¤„ç†æ—¶é•¿: 2.00å°æ—¶" in summary
    
    def test_cleanup_resources(self):
        """æµ‹è¯•èµ„æºæ¸…ç†"""
        # æ¨¡æ‹ŸæŠ“å–å™¨
        mock_scrapers = [Mock(), Mock(), Mock()]
        self.selector._active_scrapers = mock_scrapers
        
        # æ‰§è¡Œæ¸…ç†
        self.selector.cleanup_resources()
        
        # éªŒè¯æ‰€æœ‰æŠ“å–å™¨éƒ½è¢«å…³é—­
        for scraper in mock_scrapers:
            scraper.close.assert_called_once()
        
        # éªŒè¯æ´»è·ƒæŠ“å–å™¨åˆ—è¡¨è¢«æ¸…ç©º
        assert len(self.selector._active_scrapers) == 0
    
    def test_context_manager(self):
        """æµ‹è¯•ä¸Šä¸‹æ–‡ç®¡ç†å™¨"""
        with patch.object(self.selector, 'cleanup_resources') as mock_cleanup:
            with self.selector as selector:
                assert selector is self.selector
            
            # éªŒè¯é€€å‡ºæ—¶è°ƒç”¨äº†cleanup_resources
            mock_cleanup.assert_called_once()
    
    def test_error_recovery(self):
        """æµ‹è¯•é”™è¯¯æ¢å¤æœºåˆ¶"""
        # åˆ›å»ºæµ‹è¯•Excelæ–‡ä»¶
        store_data = [
            ['STORE001', '', 'å¾…å¤„ç†', 600000, 300, ''],
            ['STORE002', '', 'å¾…å¤„ç†', 800000, 400, '']
        ]
        
        excel_file = self.create_test_excel_file(store_data)
        
        try:
            # æ¨¡æ‹Ÿç¬¬ä¸€ä¸ªåº—é“ºå¤„ç†å¤±è´¥ï¼Œç¬¬äºŒä¸ªæˆåŠŸ
            def side_effect(store):
                if store.store_id == 'STORE001':
                    raise Exception("å¤„ç†å¤±è´¥")
                else:
                    return StoreAnalysisResult(
                        store_info=store,
                        products=[],
                        total_products=0,
                        profitable_products=0,
                        profit_rate=0.0,
                        is_good_store=False
                    )
            
            with patch.object(self.selector, 'process_single_store', side_effect=side_effect):
                result = self.selector.process_stores_from_excel(excel_file)
                
                # éªŒè¯é”™è¯¯æ¢å¤
                assert result.total_stores == 2
                assert result.processed_stores == 1  # åªæœ‰ä¸€ä¸ªæˆåŠŸ
                assert result.failed_stores == 1     # ä¸€ä¸ªå¤±è´¥
                
        finally:
            os.unlink(excel_file)


class TestGoodStoreSelectorIntegration:
    """æµ‹è¯•å¥½åº—ç­›é€‰ç³»ç»Ÿçš„é›†æˆåœºæ™¯"""
    
    def test_complete_workflow_simulation(self):
        """æµ‹è¯•å®Œæ•´å·¥ä½œæµç¨‹æ¨¡æ‹Ÿ"""
        config = GoodStoreSelectorConfig()
        config.store_filter.max_products_to_check = 5  # é™åˆ¶å•†å“æ•°é‡ä»¥åŠ å¿«æµ‹è¯•
        
        selector = GoodStoreSelector(config)
        
        # åˆ›å»ºæµ‹è¯•Excelæ–‡ä»¶
        store_data = [
            ['STORE001', '', 'å¾…å¤„ç†', 600000, 300, '']
        ]
        
        wb = Workbook()
        ws = wb.active
        
        # è®¾ç½®è¡¨å¤´
        headers = ['åº—é“ºID', 'æ˜¯å¦ä¸ºå¥½åº—', 'çŠ¶æ€', 'é”€å”®é¢', 'è®¢å•æ•°', 'è¯„ä¼°æ—¶é—´']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # æ·»åŠ æ•°æ®
        for row, data in enumerate(store_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # ä¿å­˜æ–‡ä»¶
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        try:
            # æ¨¡æ‹Ÿæ‰€æœ‰å¤–éƒ¨ä¾èµ–
            with patch('apps.xuanping.common.good_store_selector.SeerfarScraper') as mock_seerfar_class, \
                 patch('apps.xuanping.common.good_store_selector.OzonScraper') as mock_ozon_class, \
                 patch('apps.xuanping.common.good_store_selector.ErpPluginScraper') as mock_erp_class:
                
                # è®¾ç½®æ¨¡æ‹ŸæŠ“å–å™¨
                mock_seerfar = mock_seerfar_class.return_value
                mock_ozon = mock_ozon_class.return_value
                mock_erp = mock_erp_class.return_value
                
                # æ¨¡æ‹Ÿé”€å”®æ•°æ®æŠ“å–ï¼ˆæ»¡è¶³åˆç­›æ¡ä»¶ï¼‰
                mock_seerfar.scrape_store_sales_data.return_value = {
                    'sales_30days': 600000.0,
                    'orders_30days': 300,
                    'success': True
                }
                
                # æ¨¡æ‹Ÿå•†å“åˆ—è¡¨æŠ“å–
                mock_seerfar.scrape_store_products.return_value = [
                    ProductInfo(
                        product_id='PROD001',
                        title='æµ‹è¯•å•†å“',
                        price_rub=2000.0,
                        image_url='http://example.com/image.jpg'
                    )
                ]
                
                # æ¨¡æ‹Ÿä»·æ ¼æŠ“å–
                mock_ozon.scrape_product_prices.return_value = {
                    'green_price_rub': 1800.0,
                    'black_price_rub': 2000.0,
                    'success': True
                }
                
                # æ¨¡æ‹ŸERPæ•°æ®æŠ“å–
                mock_erp.scrape_product_attributes.return_value = {
                    'commission_rate': 12.0,
                    'weight': 500.0,
                    'success': True
                }
                
                # æ‰§è¡Œå®Œæ•´æµç¨‹
                with selector:
                    result = selector.process_stores_from_excel(temp_file.name)
                
                # éªŒè¯ç»“æœ
                assert isinstance(result, BatchProcessingResult)
                assert result.total_stores == 1
                assert result.processed_stores == 1
                
                # éªŒè¯å„ä¸ªæŠ“å–å™¨éƒ½è¢«è°ƒç”¨
                mock_seerfar.scrape_store_sales_data.assert_called()
                mock_seerfar.scrape_store_products.assert_called()
                
        finally:
            os.unlink(temp_file.name)
    
    def test_performance_monitoring(self):
        """æµ‹è¯•æ€§èƒ½ç›‘æ§"""
        config = GoodStoreSelectorConfig()
        selector = GoodStoreSelector(config)
        
        # åˆ›å»ºæµ‹è¯•æ•°æ®
        store = StoreInfo(store_id='PERF_TEST_STORE')
        
        # æ¨¡æ‹Ÿå¤„ç†è¿‡ç¨‹
        with patch.object(selector, '_scrape_store_sales_data') as mock_sales, \
             patch.object(selector, '_validate_store_filter_conditions') as mock_validate:
            
            mock_sales.return_value = {'sales_30days': 600000.0, 'orders_30days': 300}
            mock_validate.return_value = False  # ä¸æ»¡è¶³æ¡ä»¶ï¼Œå¿«é€Ÿç»“æŸ
            
            import time
            start_time = time.time()
            
            result = selector.process_single_store(store)
            
            end_time = time.time()
            processing_time = end_time - start_time
            
            # éªŒè¯æ€§èƒ½ï¼ˆå¤„ç†æ—¶é—´åº”è¯¥å¾ˆçŸ­ï¼Œå› ä¸ºåœ¨åˆç­›é˜¶æ®µå°±ç»“æŸäº†ï¼‰
            assert processing_time < 1.0  # åº”è¯¥åœ¨1ç§’å†…å®Œæˆ
            assert isinstance(result, StoreAnalysisResult)
    
    def test_configuration_impact(self):
        """æµ‹è¯•é…ç½®å¯¹ç³»ç»Ÿè¡Œä¸ºçš„å½±å“"""
        # æµ‹è¯•ä¸åŒçš„é…ç½®è®¾ç½®
        config1 = GoodStoreSelectorConfig()
        config1.store_filter.min_sales_30days = 1000000.0  # æ›´é«˜çš„é”€å”®é¢è¦æ±‚
        
        config2 = GoodStoreSelectorConfig()
        config2.store_filter.min_sales_30days = 100000.0   # æ›´ä½çš„é”€å”®é¢è¦æ±‚
        
        selector1 = GoodStoreSelector(config1)
        selector2 = GoodStoreSelector(config2)
        
        # æµ‹è¯•ç›¸åŒçš„é”€å”®æ•°æ®åœ¨ä¸åŒé…ç½®ä¸‹çš„ç»“æœ
        sales_data = {'sales_30days': 500000.0, 'orders_30days': 300}
        
        result1 = selector1._validate_store_filter_conditions(sales_data)
        result2 = selector2._validate_store_filter_conditions(sales_data)
        
        # éªŒè¯é…ç½®å½±å“
        assert result1 == False  # ä¸æ»¡è¶³é«˜è¦æ±‚
        assert result2 == True   # æ»¡è¶³ä½è¦æ±‚


if __name__ == "__main__":
    pytest.main([__file__, "-v"])