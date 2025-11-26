"""
完整的业务流程端到端测试

测试完整的业务流程链路：
Excel → CLI → GoodStoreSelector → Orchestrator → Scraper
"""

import unittest
import tempfile
import os
from pathlib import Path
from unittest.mock import patch, MagicMock

from cli.main import main as cli_main
from good_store_selector import GoodStoreSelector
from common.services.scraping_orchestrator import ScrapingOrchestrator
from common.scrapers.ozon_scraper import OzonScraper
from common.excel_processor import ExcelStoreProcessor
from common.models.scraping_result import ScrapingResult


class TestFullBusinessWorkflow(unittest.TestCase):
    """完整业务流程测试"""
    
    def setUp(self):
        """测试初始化"""
        # 创建临时Excel文件用于测试
        self.temp_excel = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_profit_calc = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.excel_file_path = self.temp_excel.name
        self.profit_calc_path = self.temp_profit_calc.name
        
        # 创建测试Excel内容
        self._create_test_excel()
        
    def tearDown(self):
        """测试清理"""
        try:
            os.unlink(self.excel_file_path)
            os.unlink(self.profit_calc_path)
        except:
            pass
    
    def _create_test_excel(self):
        """创建测试Excel文件"""
        from openpyxl import Workbook
        
        # 创建店铺数据Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "店铺列表"
        
        # 添加表头
        headers = ['店铺ID', '是否好店', '状态']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 添加测试数据
        test_stores = [
            ['STORE_001', '', ''],
            ['STORE_002', '', ''],
            ['STORE_003', '', 'PENDING']
        ]
        
        for row, store_data in enumerate(test_stores, 2):
            for col, value in enumerate(store_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(self.excel_file_path)
        
        # 创建利润计算器Excel（简化版）
        profit_wb = Workbook()
        profit_ws = profit_wb.active
        profit_ws.title = "利润计算"
        
        # 添加表头
        profit_headers = ['黑标价格', '绿标价格', '佣金率', '重量', '利润金额', '利润率']
        for col, header in enumerate(profit_headers, 1):
            profit_ws.cell(row=1, column=col, value=header)
        
        # 添加示例数据
        profit_ws.cell(row=2, column=1, value=1000)  # 黑标价格
        profit_ws.cell(row=2, column=2, value=800)   # 绿标价格
        profit_ws.cell(row=2, column=3, value=15)    # 佣金率
        profit_ws.cell(row=2, column=4, value=500)   # 重量
        
        profit_wb.save(self.profit_calc_path)
    
    @patch('common.scrapers.global_browser_singleton.get_global_browser_service')
    @patch('common.scrapers.ozon_scraper.OzonScraper')
    @patch('common.scrapers.seerfar_scraper.SeerfarScraper')
    def test_excel_to_selector_workflow(self, mock_seerfar_scraper_class, mock_ozon_scraper_class, mock_get_global_browser_service):
        """测试从Excel到店铺选择器的完整流程"""
        # 设置Mock
        mock_browser_service = MagicMock()
        mock_browser_service.navigate_to_sync.return_value = True
        mock_browser_service.evaluate_sync.return_value = "<html></html>"
        mock_get_global_browser_service.return_value = mock_browser_service
        
        # Mock OzonScraper
        mock_ozon_scraper = MagicMock()
        mock_ozon_scraper_class.return_value = mock_ozon_scraper
        mock_ozon_scraper.scrape_product_basics.return_value = ScrapingResult(
            success=True,
            data={'green_price': 1000.0, 'black_price': 1200.0},
            execution_time=0.1
        )
        
        # Mock SeerfarScraper
        mock_seerfar_scraper = MagicMock()
        mock_seerfar_scraper_class.return_value = mock_seerfar_scraper
        mock_seerfar_scraper.scrape_store_sales_data.return_value = ScrapingResult(
            success=True,
            data={
                'sales_data': {
                    'sold_30days': 500000.0,
                    'sold_count_30days': 250,
                    'daily_avg_sold': 8333.33
                },
                'products': [
                    {'product_id': 'P001', 'image_url': 'http://test.com/image1.jpg'},
                    {'product_id': 'P002', 'image_url': 'http://test.com/image2.jpg'}
                ]
            },
            execution_time=0.2
        )
        
        # 创建协调器
        orchestrator = ScrapingOrchestrator()
        
        # 测试Excel处理器
        excel_processor = ExcelStoreProcessor(self.excel_file_path)
        stores = excel_processor.read_store_data()
        
        # 验证Excel读取
        self.assertGreater(len(stores), 0)
        self.assertEqual(stores[0].store_id, 'STORE_001')
        
        # 测试店铺筛选流程
        pending_stores = excel_processor.filter_pending_stores(stores)
        self.assertGreater(len(pending_stores), 0)
        
        # 测试单个店铺处理
        if pending_stores:
            first_store = pending_stores[0]
            
            # 使用协调器处理店铺
            result = orchestrator.scrape_with_orchestration(
                mode='store_analysis',
                url='',
                store_id=first_store.store_id
            )
            
            # 验证结果
            self.assertTrue(result.success)
            self.assertIn('sales_data', result.data)
            self.assertIn('products', result.data)
    
    @patch('cli.main.GoodStoreSelector')
    def test_cli_to_selector_integration(self, mock_selector_class):
        """测试CLI到店铺选择器的集成"""
        # Mock选择器
        mock_selector = MagicMock()
        mock_selector_class.return_value = mock_selector
        
        mock_result = MagicMock()
        mock_result.total_stores = 3
        mock_result.processed_stores = 3
        mock_result.good_stores = 1
        mock_result.failed_stores = 0
        mock_result.processing_time = 1.5
        mock_selector.process_stores.return_value = mock_result
        
        # 模拟CLI调用
        test_args = [
            'start',
            self.excel_file_path,
            self.profit_calc_path,
            '--select-shops',
            '--dryrun'
        ]
        
        with patch('sys.argv', ['cli_main'] + test_args):
            try:
                cli_main()
            except SystemExit:
                pass  # CLI正常退出
        
        # 验证调用
        mock_selector_class.assert_called_once()
        mock_selector.process_stores.assert_called_once()
    
    def test_orchestrator_to_scraper_integration(self):
        """测试协调器到Scraper的集成"""
        with patch('common.scrapers.global_browser_singleton.get_global_browser_service') as mock_get_global_browser_service:
            # 设置Mock浏览器服务
            mock_browser_service = MagicMock()
            mock_browser_service.navigate_to_sync.return_value = True
            mock_browser_service.evaluate_sync.return_value = "<html></html>"
            mock_get_global_browser_service.return_value = mock_browser_service
            
            # 创建协调器
            orchestrator = ScrapingOrchestrator()
            
            # 验证Scraper初始化
            self.assertIsNotNone(orchestrator.ozon_scraper)
            self.assertIsNotNone(orchestrator.seerfar_scraper)
            self.assertIsNotNone(orchestrator.competitor_scraper)
            self.assertIsNotNone(orchestrator.erp_plugin_scraper)
            
            # 测试健康检查
            health_status = orchestrator.health_check()
            self.assertEqual(health_status['orchestrator'], 'healthy')
            self.assertEqual(health_status['browser_service'], 'healthy')
            
            # 测试Scraper获取
            ozon_scraper = orchestrator.get_scraper_by_type('ozon')
            self.assertEqual(ozon_scraper, orchestrator.ozon_scraper)

if __name__ == '__main__':
    unittest.main()
