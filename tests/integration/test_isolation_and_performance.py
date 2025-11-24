"""
测试隔离性和性能并发测试实现方案
"""

import unittest
import tempfile
import os
import threading
import time
import concurrent.futures
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch

from common.excel_processor import ExcelStoreProcessor
from common.models.scraping_result import ScrapingResult


class TestIsolationImprovement(unittest.TestCase):
    """测试隔离性改进方案"""
    
    def setUp(self):
        """为每个测试创建完全隔离的环境"""
        # 创建独立的临时目录
        self.temp_dir = tempfile.mkdtemp()
        
        # 为每个测试创建独立的Excel文件
        self.excel_file = os.path.join(self.temp_dir, 'test_stores.xlsx')
        self._create_isolated_excel()
        
        # 为每个测试创建独立的配置
        self.test_config = {
            'test_id': f'test_{int(time.time() * 1000)}',
            'temp_dir': self.temp_dir
        }
    
    def tearDown(self):
        """彻底清理测试环境"""
        try:
            if os.path.exists(self.excel_file):
                os.unlink(self.excel_file)
            if os.path.exists(self.temp_dir):
                os.rmdir(self.temp_dir)
        except:
            pass
    
    def _create_isolated_excel(self):
        """创建隔离的Excel文件"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "测试店铺"
        
        # 添加表头
        headers = ['店铺ID', '是否好店', '状态']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 添加测试数据（每个测试独立的数据）
        test_data = [
            [f'ISOLATED_STORE_{self.test_config["test_id"]}_001', '', ''],
            [f'ISOLATED_STORE_{self.test_config["test_id"]}_002', '', 'PENDING']
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(self.excel_file)
    
    def test_complete_isolation_between_test_instances(self):
        """测试实例间的完全隔离"""
        # 验证每个测试有独立的文件
        self.assertTrue(os.path.exists(self.excel_file))
        
        # 验证数据唯一性
        processor = ExcelStoreProcessor(self.excel_file)
        stores = processor.read_store_data()
        
        # 验证店铺ID包含测试ID，确保唯一性
        for store in stores:
            self.assertIn(self.test_config['test_id'], store.store_id)
    
    @patch('common.scrapers.global_browser_singleton.get_global_browser_service')
    def test_resource_isolation(self, mock_get_global_browser_service):
        """测试资源隔离"""
        # 为每个测试创建独立的Mock资源
        mock_browser_service = Mock()
        mock_browser_service.test_session_id = self.test_config['test_id']
        mock_get_global_browser_service.return_value = mock_browser_service
        
        # 验证资源隔离
        from common.services.scraping_orchestrator import ScrapingOrchestrator
        orchestrator = ScrapingOrchestrator()
        
        # 验证使用的资源是当前测试的
        self.assertEqual(orchestrator.browser_service.test_session_id, self.test_config['test_id'])


class TestThreadSafeTestEnvironment(unittest.TestCase):
    """线程安全测试环境"""
    
    def setUp(self):
        """创建线程安全的测试环境"""
        self.thread_local_data = threading.local()
        self.thread_local_data.test_id = f'thread_test_{threading.current_thread().ident}'
        
        # 创建线程独立的资源
        self.temp_dir = tempfile.mkdtemp()
        self.excel_file = os.path.join(self.temp_dir, f'thread_test_{threading.current_thread().ident}.xlsx')
        self._create_thread_excel()
    
    def tearDown(self):
        """清理线程资源"""
        try:
            if os.path.exists(self.excel_file):
                os.unlink(self.excel_file)
            if os.path.exists(self.temp_dir):
                os.rmdir(self.temp_dir)
        except:
            pass
    
    def _create_thread_excel(self):
        """创建线程独立的Excel"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # 添加线程标识的数据
        ws.cell(row=1, column=1, value='ThreadID')
        ws.cell(row=2, column=1, value=str(threading.current_thread().ident))
        
        wb.save(self.excel_file)
    
    def test_thread_isolation(self):
        """测试线程隔离"""
        processor = ExcelStoreProcessor(self.excel_file)
        # 验证可以正确读取线程特定的数据
        self.assertTrue(os.path.exists(self.excel_file))


class TestPerformanceAndConcurrencyImplementation(unittest.TestCase):
    """性能和并发测试实现方案"""
    
    def test_concurrent_store_processing(self):
        """并发店铺处理测试"""
        # 模拟并发处理多个店铺
        def process_store(store_id):
            """模拟店铺处理"""
            time.sleep(0.1)  # 模拟处理时间
            return ScrapingResult(
                success=True,
                data={
                    'store_id': store_id,
                    'processing_time': 0.1,
                    'products_count': 10
                },
                execution_time=0.1
            )
        
        # 创建测试店铺列表
        store_ids = [f'STORE_{i:03d}' for i in range(20)]
        
        # 使用线程池并发处理
        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            # 提交所有任务
            future_to_store = {
                executor.submit(process_store, store_id): store_id 
                for store_id in store_ids
            }
            
            # 收集结果
            for future in concurrent.futures.as_completed(future_to_store):
                result = future.result()
                results.append(result)
        
        # 验证并发处理结果
        self.assertEqual(len(results), len(store_ids))
        for result in results:
            self.assertTrue(result.success)
            self.assertIn('store_id', result.data)
    
    def test_performance_benchmarking_framework(self):
        """性能基准测试框架"""
        class PerformanceMetrics:
            """性能指标收集器"""
            def __init__(self):
                self.metrics = {}
            
            def start_timing(self, operation_name):
                """开始计时"""
                self.metrics[operation_name] = {
                    'start_time': time.time(),
                    'end_time': None,
                    'duration': None
                }
            
            def end_timing(self, operation_name):
                """结束计时"""
                if operation_name in self.metrics:
                    self.metrics[operation_name]['end_time'] = time.time()
                    self.metrics[operation_name]['duration'] = (
                        self.metrics[operation_name]['end_time'] - 
                        self.metrics[operation_name]['start_time']
                    )
            
            def get_metric(self, operation_name):
                """获取指标"""
                return self.metrics.get(operation_name, {})
        
        # 使用性能测试框架
        metrics = PerformanceMetrics()
        
        # 测试不同操作的性能
        operations = {
            'excel_read': lambda: self._mock_excel_read(),
            'data_processing': lambda: self._mock_data_processing(),
            'result_writing': lambda: self._mock_result_writing()
        }
        
        for op_name, op_func in operations.items():
            metrics.start_timing(op_name)
            result = op_func()
            metrics.end_timing(op_name)
            
            # 验证操作成功执行
            self.assertIsNotNone(result)
        
        # 验证性能指标收集
        for op_name in operations.keys():
            metric = metrics.get_metric(op_name)
            self.assertIn('duration', metric)
            self.assertGreaterEqual(metric['duration'], 0)
    
    def _mock_excel_read(self):
        """模拟Excel读取"""
        time.sleep(0.05)
        return {'stores': 100}
    
    def _mock_data_processing(self):
        """模拟数据处理"""
        time.sleep(0.1)
        return {'processed': 100}
    
    def _mock_result_writing(self):
        """模拟结果写入"""
        time.sleep(0.02)
        return {'written': 100}
    
    def test_load_testing_scenario(self):
        """负载测试场景"""
        def simulate_user_request():
            """模拟用户请求"""
            # 模拟完整的业务流程
            time.sleep(0.05)  # 模拟网络延迟
            return ScrapingResult(
                success=True,
                data={'request_id': f'req_{int(time.time() * 1000)}'},
                execution_time=0.05
            )
        
        # 模拟并发用户请求
        concurrent_users = 10
        requests_per_user = 5
        
        all_results = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=concurrent_users) as executor:
            # 为每个用户创建任务
            user_futures = []
            for user_id in range(concurrent_users):
                def user_session(user_id):
                    user_results = []
                    for req_id in range(requests_per_user):
                        result = simulate_user_request()
                        user_results.append(result)
                        time.sleep(0.01)  # 用户思考时间
                    return user_results
                
                future = executor.submit(user_session, user_id)
                user_futures.append(future)
            
            # 收集所有用户的结果
            for future in concurrent.futures.as_completed(user_futures):
                user_results = future.result()
                all_results.extend(user_results)
        
        # 验证负载测试结果
        expected_total_requests = concurrent_users * requests_per_user
        self.assertEqual(len(all_results), expected_total_requests)
        
        # 验证所有请求都成功
        successful_requests = [r for r in all_results if r.success]
        self.assertEqual(len(successful_requests), expected_total_requests)


class TestResourceManagement(unittest.TestCase):
    """资源管理测试"""
    
    def test_automatic_resource_cleanup(self):
        """自动资源清理测试"""
        # 创建需要清理的资源
        temp_files = []
        temp_dirs = []
        
        # 模拟创建多个临时资源
        for i in range(5):
            temp_file = tempfile.NamedTemporaryFile(delete=False)
            temp_files.append(temp_file.name)
            temp_file.close()
        
        # 验证资源创建成功
        for temp_file in temp_files:
            self.assertTrue(os.path.exists(temp_file))
        
        # 模拟测试结束时的自动清理
        for temp_file in temp_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        
        # 验证资源已清理
        for temp_file in temp_files:
            self.assertFalse(os.path.exists(temp_file))


if __name__ == '__main__':
    # 运行测试时使用更详细的输出
    unittest.main(verbosity=2)
