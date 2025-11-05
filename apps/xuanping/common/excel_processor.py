"""
Excel数据处理模块

基于现有excel_calculator.py扩展，实现店铺数据的读取、验证和更新功能。
遵循规格要求，支持Excel文件的标准化处理和状态管理。
"""

import logging
from typing import List, Optional, Dict, Any, Tuple
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    ExcelStoreData, StoreInfo, StoreStatus, GoodStoreFlag,
    ExcelProcessingError, DataValidationError
)
from .config import GoodStoreSelectorConfig, get_config
from .excel_calculator import ExcelProfitCalculator, ProfitCalculatorResult


class ExcelStoreProcessor:
    """Excel店铺数据处理器"""
    
    def __init__(self, excel_file_path: str, config: Optional[GoodStoreSelectorConfig] = None):
        """
        初始化Excel处理器
        
        Args:
            excel_file_path: Excel文件路径
            config: 配置对象，如果为None则使用全局配置
        """
        self.config = config or get_config()
        self.excel_file_path = Path(excel_file_path)
        self.logger = logging.getLogger(f"{__name__}.ExcelStoreProcessor")
        
        # 验证文件存在性
        if not self.excel_file_path.exists():
            raise ExcelProcessingError(f"Excel文件不存在: {self.excel_file_path}")
        
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self._load_workbook()
    
    def _load_workbook(self):
        """加载Excel工作簿"""
        try:
            self.workbook = load_workbook(self.excel_file_path)
            
            # 获取第一个工作表
            if not self.workbook.worksheets:
                raise ExcelProcessingError("Excel文件中没有工作表")
            
            self.worksheet = self.workbook.active
            self.logger.info(f"成功加载Excel文件: {self.excel_file_path}")
            
        except Exception as e:
            raise ExcelProcessingError(f"加载Excel文件失败: {e}")
    
    def read_store_data(self) -> List[ExcelStoreData]:
        """
        读取Excel中的店铺数据
        
        Returns:
            List[ExcelStoreData]: 店铺数据列表
        """
        if not self.worksheet:
            raise ExcelProcessingError("工作表未加载")
        
        store_data_list = []
        max_rows = min(self.worksheet.max_row, self.config.excel.max_rows_to_process)
        
        # 从第2行开始读取（假设第1行是表头）
        for row_idx in range(2, max_rows + 1):
            try:
                # 读取店铺ID
                store_id_cell = self.worksheet[f"{self.config.excel.store_id_column}{row_idx}"]
                store_id = str(store_id_cell.value).strip() if store_id_cell.value else ""
                
                # 跳过空行
                if not store_id and self.config.excel.skip_empty_rows:
                    continue
                
                if not store_id:
                    raise DataValidationError(f"第{row_idx}行店铺ID为空")
                
                # 读取是否为好店
                good_store_cell = self.worksheet[f"{self.config.excel.good_store_column}{row_idx}"]
                good_store_value = str(good_store_cell.value).strip() if good_store_cell.value else ""
                
                try:
                    is_good_store = GoodStoreFlag(good_store_value) if good_store_value else GoodStoreFlag.EMPTY
                except ValueError:
                    self.logger.warning(f"第{row_idx}行好店标记值无效: {good_store_value}，设为空")
                    is_good_store = GoodStoreFlag.EMPTY
                
                # 读取状态
                status_cell = self.worksheet[f"{self.config.excel.status_column}{row_idx}"]
                status_value = str(status_cell.value).strip() if status_cell.value else ""
                
                try:
                    status = StoreStatus(status_value) if status_value else StoreStatus.EMPTY
                except ValueError:
                    self.logger.warning(f"第{row_idx}行状态值无效: {status_value}，设为空")
                    status = StoreStatus.EMPTY
                
                # 创建数据对象
                store_data = ExcelStoreData(
                    row_index=row_idx,
                    store_id=store_id,
                    is_good_store=is_good_store,
                    status=status
                )
                
                store_data_list.append(store_data)
                
            except Exception as e:
                self.logger.error(f"读取第{row_idx}行数据失败: {e}")
                if not self.config.excel.skip_empty_rows:
                    raise ExcelProcessingError(f"读取第{row_idx}行数据失败: {e}")
        
        self.logger.info(f"成功读取{len(store_data_list)}条店铺数据")
        return store_data_list
    
    def filter_pending_stores(self, store_data_list: List[ExcelStoreData]) -> List[ExcelStoreData]:
        """
        筛选待处理的店铺
        
        Args:
            store_data_list: 店铺数据列表
            
        Returns:
            List[ExcelStoreData]: 待处理的店铺数据列表
        """
        pending_stores = [
            store for store in store_data_list 
            if store.status == StoreStatus.PENDING
        ]
        
        self.logger.info(f"筛选出{len(pending_stores)}个待处理店铺（总共{len(store_data_list)}个）")
        return pending_stores
    
    def update_store_status(self, store_data: ExcelStoreData, 
                           is_good_store: GoodStoreFlag, 
                           status: StoreStatus = StoreStatus.PROCESSED):
        """
        更新店铺状态
        
        Args:
            store_data: 店铺数据
            is_good_store: 是否为好店
            status: 处理状态
        """
        if not self.worksheet:
            raise ExcelProcessingError("工作表未加载")
        
        try:
            # 更新是否为好店
            good_store_cell = f"{self.config.excel.good_store_column}{store_data.row_index}"
            self.worksheet[good_store_cell] = is_good_store.value
            
            # 更新状态
            status_cell = f"{self.config.excel.status_column}{store_data.row_index}"
            self.worksheet[status_cell] = status.value
            
            self.logger.debug(f"更新店铺{store_data.store_id}状态: 好店={is_good_store.value}, 状态={status.value}")
            
        except Exception as e:
            raise ExcelProcessingError(f"更新店铺{store_data.store_id}状态失败: {e}")
    
    def batch_update_stores(self, updates: List[Tuple[ExcelStoreData, GoodStoreFlag, StoreStatus]]):
        """
        批量更新店铺状态
        
        Args:
            updates: 更新列表，每个元素为(store_data, is_good_store, status)
        """
        for store_data, is_good_store, status in updates:
            self.update_store_status(store_data, is_good_store, status)
        
        self.logger.info(f"批量更新{len(updates)}个店铺状态")
    
    def save_changes(self):
        """保存Excel文件更改"""
        if not self.workbook:
            raise ExcelProcessingError("工作簿未加载")
        
        try:
            if not self.config.dry_run:
                self.workbook.save(self.excel_file_path)
                self.logger.info(f"Excel文件已保存: {self.excel_file_path}")
            else:
                self.logger.info("干运行模式，跳过保存Excel文件")
                
        except Exception as e:
            raise ExcelProcessingError(f"保存Excel文件失败: {e}")
    
    def close(self):
        """关闭Excel文件"""
        if self.workbook:
            try:
                self.workbook.close()
                self.logger.info("Excel文件已关闭")
            except Exception as e:
                self.logger.warning(f"关闭Excel文件时出现警告: {e}")
            finally:
                self.workbook = None
                self.worksheet = None
    
    def validate_excel_format(self) -> bool:
        """
        验证Excel文件格式
        
        Returns:
            bool: 格式是否正确
        """
        if not self.worksheet:
            return False
        
        try:
            # 检查是否有足够的列
            required_columns = [
                self.config.excel.store_id_column,
                self.config.excel.good_store_column,
                self.config.excel.status_column
            ]
            
            # 检查第一行是否有数据（表头）
            for col in required_columns:
                cell_value = self.worksheet[f"{col}1"].value
                if cell_value is None:
                    self.logger.warning(f"列{col}的表头为空")
            
            # 检查是否有数据行
            if self.worksheet.max_row < 2:
                self.logger.warning("Excel文件没有数据行")
                return False
            
            return True
            
        except Exception as e:
            self.logger.error(f"验证Excel格式失败: {e}")
            return False
    
    def get_statistics(self) -> Dict[str, Any]:
        """
        获取Excel数据统计信息
        
        Returns:
            Dict[str, Any]: 统计信息
        """
        if not self.worksheet:
            return {}
        
        try:
            store_data_list = self.read_store_data()
            
            total_stores = len(store_data_list)
            pending_stores = len([s for s in store_data_list if s.status == StoreStatus.PENDING])
            processed_stores = len([s for s in store_data_list if s.status == StoreStatus.PROCESSED])
            good_stores = len([s for s in store_data_list if s.is_good_store == GoodStoreFlag.YES])
            
            return {
                'total_stores': total_stores,
                'pending_stores': pending_stores,
                'processed_stores': processed_stores,
                'good_stores': good_stores,
                'file_path': str(self.excel_file_path),
                'max_row': self.worksheet.max_row,
                'max_column': self.worksheet.max_column
            }
            
        except Exception as e:
            self.logger.error(f"获取统计信息失败: {e}")
            return {'error': str(e)}


class ExcelProfitProcessor:
    """Excel利润计算处理器，集成现有的ExcelProfitCalculator"""
    
    def __init__(self, profit_calculator_path: str, config: Optional[GoodStoreSelectorConfig] = None):
        """
        初始化利润计算处理器
        
        Args:
            profit_calculator_path: 利润计算器Excel文件路径
            config: 配置对象
        """
        self.config = config or get_config()
        self.profit_calculator_path = Path(profit_calculator_path)
        self.logger = logging.getLogger(f"{__name__}.ExcelProfitProcessor")
        
        # 初始化利润计算器
        self.calculator = ExcelProfitCalculator(self.profit_calculator_path)
    
    def calculate_product_profit(self, black_price: float, green_price: float, 
                               commission_rate: float, weight: float) -> ProfitCalculatorResult:
        """
        计算商品利润
        
        Args:
            black_price: 黑标价格
            green_price: 绿标价格
            commission_rate: 佣金率
            weight: 重量
            
        Returns:
            ProfitCalculatorResult: 计算结果
        """
        try:
            result = self.calculator.calculate_profit(
                black_price=black_price,
                green_price=green_price,
                commission_rate=commission_rate,
                weight=weight
            )
            
            self.logger.debug(f"利润计算完成: 利润={result.profit_amount:.2f}, 利润率={result.profit_rate:.2f}%")
            return result
            
        except Exception as e:
            self.logger.error(f"利润计算失败: {e}")
            raise PriceCalculationError(f"利润计算失败: {e}")
    
    def batch_calculate_profits(self, calculations: List[Dict[str, float]]) -> List[ProfitCalculatorResult]:
        """
        批量计算利润
        
        Args:
            calculations: 计算参数列表
            
        Returns:
            List[ProfitCalculatorResult]: 计算结果列表
        """
        results = []
        
        for i, calc_params in enumerate(calculations):
            try:
                result = self.calculate_product_profit(
                    black_price=calc_params['black_price'],
                    green_price=calc_params['green_price'],
                    commission_rate=calc_params['commission_rate'],
                    weight=calc_params['weight']
                )
                results.append(result)
                
            except Exception as e:
                self.logger.error(f"第{i+1}个商品利润计算失败: {e}")
                # 创建错误结果
                error_result = ProfitCalculatorResult(
                    profit_amount=0.0,
                    profit_rate=0.0,
                    is_loss=True,
                    input_summary=calc_params,
                    calculation_time=0.0,
                    log_info={'error': str(e), 'status': 'error'}
                )
                results.append(error_result)
        
        self.logger.info(f"批量计算完成，共{len(calculations)}个商品")
        return results
    
    def close(self):
        """关闭利润计算器"""
        if self.calculator:
            self.calculator.close()


# 便捷函数

def read_stores_from_excel(excel_file_path: str, 
                          config: Optional[GoodStoreSelectorConfig] = None) -> List[ExcelStoreData]:
    """
    从Excel文件读取店铺数据的便捷函数
    
    Args:
        excel_file_path: Excel文件路径
        config: 配置对象
        
    Returns:
        List[ExcelStoreData]: 店铺数据列表
    """
    processor = None
    try:
        processor = ExcelStoreProcessor(excel_file_path, config)
        return processor.read_store_data()
    finally:
        if processor:
            processor.close()


def get_pending_stores(excel_file_path: str, 
                      config: Optional[GoodStoreSelectorConfig] = None) -> List[ExcelStoreData]:
    """
    获取待处理店铺的便捷函数
    
    Args:
        excel_file_path: Excel文件路径
        config: 配置对象
        
    Returns:
        List[ExcelStoreData]: 待处理店铺数据列表
    """
    processor = None
    try:
        processor = ExcelStoreProcessor(excel_file_path, config)
        all_stores = processor.read_store_data()
        return processor.filter_pending_stores(all_stores)
    finally:
        if processor:
            processor.close()


def update_store_results(excel_file_path: str, 
                        updates: List[Tuple[ExcelStoreData, GoodStoreFlag, StoreStatus]],
                        config: Optional[GoodStoreSelectorConfig] = None):
    """
    更新店铺结果的便捷函数
    
    Args:
        excel_file_path: Excel文件路径
        updates: 更新列表
        config: 配置对象
    """
    processor = None
    try:
        processor = ExcelStoreProcessor(excel_file_path, config)
        processor.batch_update_stores(updates)
        processor.save_changes()
    finally:
        if processor:
            processor.close()


def validate_excel_file(excel_file_path: str, 
                       config: Optional[GoodStoreSelectorConfig] = None) -> bool:
    """
    验证Excel文件格式的便捷函数
    
    Args:
        excel_file_path: Excel文件路径
        config: 配置对象
        
    Returns:
        bool: 格式是否正确
    """
    processor = None
    try:
        processor = ExcelStoreProcessor(excel_file_path, config)
        return processor.validate_excel_format()
    except Exception:
        return False
    finally:
        if processor:
            processor.close()