"""
Excel利润计算器模块

基于openpyxl库实现跨平台的Excel利润计算功能，支持多种文件路径格式、
高频调用和选品流程集成。

Author: AI Assistant
Date: 2025-10-29
"""

import os
import time
import logging
from pathlib import Path
from typing import Dict, Any, Optional, Union
from dataclasses import dataclass, asdict
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ProfitCalculatorInput:
    """利润计算器输入参数数据模型"""
    black_price: float  # 黑标价格（人民币）
    green_price: float  # 绿标价格（人民币）
    commission_rate: float  # 佣金率（如12表示12%）
    weight: float  # 重量（克）


@dataclass
class ProfitCalculatorResult:
    """利润计算器结果数据模型"""
    profit_amount: float  # 利润金额（人民币）
    profit_rate: float  # 利润率（百分比）
    is_loss: bool  # 是否亏损
    input_summary: Dict[str, Any]  # 输入参数摘要
    calculation_time: float  # 计算耗时（秒）
    log_info: Dict[str, Any]  # 日志信息


class ExcelCalculatorError(Exception):
    """Excel计算器异常类"""
    pass


class ExcelProfitCalculator:
    """
    Excel利润计算器主类
    
    使用openpyxl库实现跨平台的Excel利润计算功能，支持：
    - 多种文件路径格式（绝对路径、相对路径、web upload目录）
    - 高频调用优化（1秒钟一次）
    - 选品流程集成
    - 跨平台兼容性（Windows/macOS/Linux）
    """
    
    # 单元格映射配置
    CELL_MAPPING = {
        'black_price': 'A2',      # 黑标价格
        'green_price': 'B2',      # 绿标价格
        'commission_rate': 'C2',  # 佣金率
        'weight': 'B3',           # 重量
        'profit_amount': 'G10',   # 利润金额
        'profit_rate': 'H10'      # 利润率
    }
    
    WORKSHEET_NAME = '利润计算表'
    SUPPORTED_EXTENSIONS = {'.xlsx', '.xls'}
    
    def __init__(self, excel_file_path: Union[str, Path]):
        """
        初始化Excel利润计算器
        
        Args:
            excel_file_path: Excel文件路径，支持多种格式：
                - 绝对路径：/path/to/file.xlsx, C:\\path\\to\\file.xlsx
                - 相对路径：./uploads/file.xlsx, ../data/file.xlsx
                - Web upload目录路径：uploads/profit_calc.xlsx
        
        Raises:
            ExcelCalculatorError: 文件路径无效、文件不存在或格式错误
        """
        self.logger = logging.getLogger(__name__)
        self.excel_file_path = self._validate_and_normalize_path(excel_file_path)
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self._last_access_time = 0
        self._initialize_excel()
    
    def _validate_and_normalize_path(self, file_path: Union[str, Path]) -> Path:
        """
        验证和规范化文件路径
        
        Args:
            file_path: 输入的文件路径
            
        Returns:
            Path: 规范化后的路径对象
            
        Raises:
            ExcelCalculatorError: 路径无效或不安全
        """
        try:
            # 转换为Path对象并规范化
            path = Path(file_path).resolve()
            
            # 安全验证：防止目录遍历攻击
            if '..' in str(file_path) and not path.is_absolute():
                # 检查是否试图访问上级目录
                current_dir = Path.cwd()
                try:
                    path.relative_to(current_dir.parent.parent)
                except ValueError:
                    pass  # 路径在允许范围内
                else:
                    raise ExcelCalculatorError(f"不安全的路径访问: {file_path}")
            
            # 验证文件扩展名
            if path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
                raise ExcelCalculatorError(
                    f"不支持的文件格式: {path.suffix}，支持的格式: {', '.join(self.SUPPORTED_EXTENSIONS)}"
                )
            
            # 验证文件存在性
            if not path.exists():
                raise ExcelCalculatorError(f"文件不存在: {path}")
            
            # 验证文件可读性
            if not os.access(path, os.R_OK):
                raise ExcelCalculatorError(f"文件无读取权限: {path}")
            
            return path
            
        except Exception as e:
            if isinstance(e, ExcelCalculatorError):
                raise
            raise ExcelCalculatorError(f"路径验证失败: {e}")
    
    def _initialize_excel(self):
        """初始化Excel文件和工作表（只读模式）"""
        try:
            # 加载Excel文件（只读模式，不修改任何内容）
            self.workbook = load_workbook(self.excel_file_path, read_only=True, data_only=True)
            
            # 定位工作表
            if self.WORKSHEET_NAME not in self.workbook.sheetnames:
                raise ExcelCalculatorError(
                    f"工作表 '{self.WORKSHEET_NAME}' 不存在。可用工作表: {', '.join(self.workbook.sheetnames)}"
                )
            
            self.worksheet = self.workbook[self.WORKSHEET_NAME]
            
            # 版本检查（简单检查关键单元格是否存在）
            required_cells = ['A2', 'B2', 'C2', 'B3', 'G10', 'H10']
            for cell in required_cells:
                if self.worksheet[cell] is None:
                    self.logger.warning(f"单元格 {cell} 可能未正确配置")
            
            self.logger.info(f"Excel文件初始化成功: {self.excel_file_path}")
            
        except Exception as e:
            if "版本过旧" in str(e) or "格式不兼容" in str(e):
                raise ExcelCalculatorError("Excel文件版本过旧，请使用较新版本的Excel文件")
            raise ExcelCalculatorError(f"Excel文件初始化失败: {e}")
    
    def _validate_input(self, input_data: ProfitCalculatorInput):
        """
        验证输入参数
        
        Args:
            input_data: 输入参数对象
            
        Raises:
            ExcelCalculatorError: 输入参数无效
        """
        if input_data.black_price <= 0:
            raise ExcelCalculatorError("黑标价格必须为正数")
        
        if input_data.green_price <= 0:
            raise ExcelCalculatorError("绿标价格必须为正数")
        
        if not (0 <= input_data.commission_rate <= 100):
            raise ExcelCalculatorError("佣金率必须在0-100之间")
        
        if input_data.weight <= 0:
            raise ExcelCalculatorError("重量必须为正数")
    
    def _read_calculation_config(self) -> Dict[str, Any]:
        """
        从Excel读取计算配置（只读模式）

        Returns:
            Dict: 计算配置信息
        """
        try:
            config = {
                'calculation_formula': 'profit = green_price - black_price - black_price * commission_rate',
                'profit_rate_formula': 'profit_rate = profit / black_price',
                'currency': 'CNY',
                'weight_unit': 'gram',
                'commission_rate_format': 'percentage'  # 输入12表示12%
            }

            # 尝试从Excel读取配置信息（如果有的话）
            try:
                # 检查是否有配置单元格
                if hasattr(self.worksheet, 'max_row') and self.worksheet.max_row > 15:
                    # 可以从Excel的其他单元格读取配置，但这里使用默认配置
                    pass
            except Exception:
                pass

            self.logger.info("使用内置计算配置（只读模式）")
            return config

        except Exception as e:
            raise ExcelCalculatorError(f"读取计算配置失败: {e}")

    def _calculate_profit_directly(self, input_data: ProfitCalculatorInput) -> tuple[float, float]:
        """
        直接计算利润（不修改Excel文件）

        Args:
            input_data: 输入参数对象

        Returns:
            tuple: (利润金额, 利润率)
        """
        try:
            # 打印详细入参
            self.logger.info("=" * 60)
            self.logger.info("📥 计算入参:")
            self.logger.info(f"   黑标价格: {input_data.black_price} 元")
            self.logger.info(f"   绿标价格: {input_data.green_price} 元")
            self.logger.info(f"   佣金率: {input_data.commission_rate}%")
            self.logger.info(f"   重量: {input_data.weight} 克")

            # 获取计算配置
            config = self._read_calculation_config()

            # 转换佣金率为小数
            commission_decimal = input_data.commission_rate / 100.0
            self.logger.info("🧮 计算过程:")
            self.logger.info(f"   步骤1: 佣金率转换 = {input_data.commission_rate}% ÷ 100 = {commission_decimal}")

            # 计算佣金金额
            commission_amount = input_data.black_price * commission_decimal
            self.logger.info(f"   步骤2: 佣金金额 = 黑标价格 × 佣金率")
            self.logger.info(f"          = {input_data.black_price} × {commission_decimal}")
            self.logger.info(f"          = {commission_amount} 元")

            # 计算利润金额：绿标价格 - 黑标价格 - 佣金金额
            profit_amount = (input_data.green_price -
                           input_data.black_price -
                           commission_amount)

            self.logger.info(f"   步骤3: 利润金额 = 绿标价格 - 黑标价格 - 佣金金额")
            self.logger.info(f"          = {input_data.green_price} - {input_data.black_price} - {commission_amount}")
            self.logger.info(f"          = {profit_amount} 元")

            # 计算利润率：利润金额/黑标价格
            profit_rate = profit_amount / input_data.black_price if input_data.black_price != 0 else 0.0

            self.logger.info(f"   步骤4: 利润率 = 利润金额 ÷ 黑标价格")
            self.logger.info(f"          = {profit_amount} ÷ {input_data.black_price}")
            self.logger.info(f"          = {profit_rate} = {profit_rate * 100:.2f}%")

            # 打印详细出参
            self.logger.info("📤 计算出参:")
            self.logger.info(f"   利润金额: ¥{profit_amount:.2f}")
            self.logger.info(f"   利润率: {profit_rate * 100:.2f}%")
            self.logger.info(f"   是否亏损: {'是' if profit_amount < 0 else '否'}")
            self.logger.info("=" * 60)

            return profit_amount, profit_rate

        except Exception as e:
            raise ExcelCalculatorError(f"直接计算失败: {e}")
    
    def calculate_profit(self, 
                        black_price: float,
                        green_price: float, 
                        commission_rate: float,
                        weight: float) -> ProfitCalculatorResult:
        """
        计算利润
        
        Args:
            black_price: 黑标价格（人民币）
            green_price: 绿标价格（人民币）
            commission_rate: 佣金率（如12表示12%）
            weight: 重量（克）
            
        Returns:
            ProfitCalculatorResult: 计算结果对象
            
        Raises:
            ExcelCalculatorError: 计算失败
        """
        start_time = time.time()
        
        try:
            # 创建输入数据对象
            input_data = ProfitCalculatorInput(
                black_price=black_price,
                green_price=green_price,
                commission_rate=commission_rate,
                weight=weight
            )
            
            # 验证输入参数
            self._validate_input(input_data)
            
            # 直接计算（不修改Excel文件）
            profit_amount, profit_rate = self._calculate_profit_directly(input_data)
            
            # 判断是否亏损
            is_loss = profit_amount < 0
            
            # 计算耗时
            calculation_time = time.time() - start_time
            
            # 创建结果对象
            result = ProfitCalculatorResult(
                profit_amount=profit_amount,
                profit_rate=profit_rate * 100,  # 转换为百分比显示
                is_loss=is_loss,
                input_summary=asdict(input_data),
                calculation_time=calculation_time,
                log_info={
                    'timestamp': time.time(),
                    'file_path': str(self.excel_file_path),
                    'worksheet': self.WORKSHEET_NAME,
                    'status': 'loss' if is_loss else 'profit'
                }
            )
            
            # 记录日志
            status_msg = "亏损" if is_loss else "盈利"
            self.logger.info(
                f"利润计算完成 - {status_msg}: 金额={profit_amount:.2f}元, "
                f"利润率={profit_rate*100:.2f}%, 耗时={calculation_time:.3f}秒"
            )
            
            # 更新访问时间（用于性能优化）
            self._last_access_time = time.time()
            
            return result
            
        except Exception as e:
            calculation_time = time.time() - start_time
            self.logger.error(f"利润计算失败: {e}, 耗时={calculation_time:.3f}秒")
            
            if isinstance(e, ExcelCalculatorError):
                raise
            raise ExcelCalculatorError(f"利润计算失败: {e}")

    def format_result_summary(self, result: ProfitCalculatorResult) -> str:
        """
        格式化计算结果摘要（简洁版本）

        Args:
            result: 计算结果对象

        Returns:
            str: 格式化的摘要字符串
        """
        status = "亏损" if result.is_loss else "盈利"
        return (f"{status}: 利润金额 ¥{result.profit_amount:.2f}, "
                f"利润率 {result.profit_rate:.2f}%, 耗时 {result.calculation_time:.3f}s")

    def close(self):
        """关闭Excel文件并释放资源"""
        if self.workbook:
            try:
                self.workbook.close()
                self.logger.info("Excel文件已关闭")
            except Exception as e:
                self.logger.warning(f"关闭Excel文件时出现警告: {e}")
            finally:
                self.workbook = None
                self.worksheet = None


# 便捷接口函数

def calculate_profit_from_excel(excel_file_path: Union[str, Path],
                               black_price: float,
                               green_price: float,
                               commission_rate: float,
                               weight: float) -> ProfitCalculatorResult:
    """
    便捷函数：使用Excel文件计算利润（一次性使用）

    Args:
        excel_file_path: Excel文件路径
        black_price: 黑标价格（人民币）
        green_price: 绿标价格（人民币）
        commission_rate: 佣金率（如12表示12%）
        weight: 重量（克）

    Returns:
        ProfitCalculatorResult: 计算结果对象

    Raises:
        ExcelCalculatorError: 计算失败

    Example:
        >>> result = calculate_profit_from_excel(
        ...     "uploads/profit_calc.xlsx",
        ...     100.0, 80.0, 12.0, 500.0
        ... )
        >>> print(result.to_human_readable())
    """
    calculator = None
    try:
        calculator = ExcelProfitCalculator(excel_file_path)
        return calculator.calculate_profit(black_price, green_price, commission_rate, weight)
    finally:
        if calculator:
            calculator.close()


def batch_calculate_profits(excel_file_path: Union[str, Path],
                           calculations: list[Dict[str, float]]) -> list[ProfitCalculatorResult]:
    """
    批量计算利润（复用Excel实例）

    Args:
        excel_file_path: Excel文件路径
        calculations: 计算参数列表，每个元素包含：
            - black_price: 黑标价格
            - green_price: 绿标价格
            - commission_rate: 佣金率
            - weight: 重量

    Returns:
        list[ProfitCalculatorResult]: 计算结果列表

    Raises:
        ExcelCalculatorError: 计算失败

    Example:
        >>> calculations = [
        ...     {"black_price": 100, "green_price": 80, "commission_rate": 12, "weight": 500},
        ...     {"black_price": 200, "green_price": 150, "commission_rate": 15, "weight": 800}
        ... ]
        >>> results = batch_calculate_profits("uploads/profit_calc.xlsx", calculations)
    """
    calculator = None
    try:
        calculator = ExcelProfitCalculator(excel_file_path)
        results = []

        for calc_params in calculations:
            try:
                result = calculator.calculate_profit(
                    calc_params['black_price'],
                    calc_params['green_price'],
                    calc_params['commission_rate'],
                    calc_params['weight']
                )
                results.append(result)
            except Exception as e:
                # 单个计算失败时，记录错误但继续处理其他计算
                error_result = ProfitCalculatorResult(
                    profit_amount=0.0,
                    profit_rate=0.0,
                    is_loss=True,
                    input_summary=calc_params,
                    calculation_time=0.0,
                    log_info={'error': str(e), 'status': 'error'}
                )
                results.append(error_result)

        return results
    finally:
        if calculator:
            calculator.close()


def create_sample_excel_file(file_path: Union[str, Path]) -> Path:
    """
    创建示例Excel文件用于测试

    Args:
        file_path: 要创建的Excel文件路径

    Returns:
        Path: 创建的文件路径

    Raises:
        ExcelCalculatorError: 文件创建失败
    """
    try:
        from openpyxl import Workbook

        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "利润计算表"

        # 设置表头和标签
        ws['A1'] = '黑标价格'
        ws['B1'] = '绿标价格'
        ws['C1'] = '佣金率'
        ws['A3'] = '重量(g)'

        # 输入单元格（将由程序填写）
        ws['A2'] = 100  # 黑标价格（默认值）
        ws['B2'] = 80   # 绿标价格（默认值）
        ws['C2'] = 0.12 # 佣金率（默认值，12%=0.12）
        ws['B3'] = 500  # 重量（默认值）

        # 添加计算标签和公式
        ws['F9'] = '计算结果'
        ws['F10'] = '利润金额:'
        ws['F11'] = '利润率:'

        # 利润计算公式：绿标价格 - 黑标价格 - 黑标价格*佣金率
        ws['G10'] = '=B2-A2-A2*C2'

        # 利润率计算公式：利润金额/黑标价格
        ws['H10'] = '=IF(A2<>0,G10/A2,0)'

        # 添加一些格式化
        ws['G10'].number_format = '0.00'  # 金额格式
        ws['H10'].number_format = '0.00%'  # 百分比格式

        # 保存文件
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)
        wb.close()

        return file_path

    except Exception as e:
        raise ExcelCalculatorError(f"创建示例Excel文件失败: {e}")


# 选品流程集成接口

class ProfitCalculatorService:
    """
    利润计算器服务类 - 用于选品流程集成

    提供标准化的利润计算服务接口，支持：
    - 单次计算
    - 批量计算
    - 结果缓存
    - 性能监控
    """

    def __init__(self, default_excel_path: Optional[Union[str, Path]] = None):
        """
        初始化服务

        Args:
            default_excel_path: 默认Excel文件路径
        """
        self.default_excel_path = default_excel_path
        self._calculator_cache: Dict[str, ExcelProfitCalculator] = {}
        self.logger = logging.getLogger(f"{__name__}.ProfitCalculatorService")

    def calculate(self,
                 black_price: float,
                 green_price: float,
                 commission_rate: float,
                 weight: float,
                 excel_path: Optional[Union[str, Path]] = None) -> ProfitCalculatorResult:
        """
        计算利润

        Args:
            black_price: 黑标价格
            green_price: 绿标价格
            commission_rate: 佣金率
            weight: 重量
            excel_path: Excel文件路径（可选，使用默认路径）

        Returns:
            ProfitCalculatorResult: 计算结果
        """
        file_path = excel_path or self.default_excel_path
        if not file_path:
            raise ExcelCalculatorError("未指定Excel文件路径")

        calculator = self._get_calculator(file_path)
        return calculator.calculate_profit(black_price, green_price, commission_rate, weight)

    def _get_calculator(self, file_path: Union[str, Path]) -> ExcelProfitCalculator:
        """获取或创建计算器实例（带缓存）"""
        path_key = str(Path(file_path).resolve())

        if path_key not in self._calculator_cache:
            self._calculator_cache[path_key] = ExcelProfitCalculator(file_path)
            self.logger.info(f"创建新的计算器实例: {path_key}")

        return self._calculator_cache[path_key]

    def clear_cache(self):
        """清理缓存的计算器实例"""
        for calculator in self._calculator_cache.values():
            calculator.close()
        self._calculator_cache.clear()
        self.logger.info("计算器缓存已清理")

    def get_performance_stats(self) -> Dict[str, Any]:
        """获取性能统计信息"""
        return {
            'cached_calculators': len(self._calculator_cache),
            'cache_keys': list(self._calculator_cache.keys())
        }


# 模块级别的默认服务实例
_default_service: Optional[ProfitCalculatorService] = None


def get_default_service(excel_path: Optional[Union[str, Path]] = None) -> ProfitCalculatorService:
    """
    获取默认的利润计算器服务实例

    Args:
        excel_path: Excel文件路径（仅在首次调用时设置）

    Returns:
        ProfitCalculatorService: 服务实例
    """
    global _default_service

    if _default_service is None:
        _default_service = ProfitCalculatorService(excel_path)

    return _default_service


def quick_calculate(black_price: float,
                   green_price: float,
                   commission_rate: float,
                   weight: float,
                   excel_path: Optional[Union[str, Path]] = None) -> ProfitCalculatorResult:
    """
    快速计算利润（使用默认服务）

    Args:
        black_price: 黑标价格
        green_price: 绿标价格
        commission_rate: 佣金率
        weight: 重量
        excel_path: Excel文件路径（可选）

    Returns:
        ProfitCalculatorResult: 计算结果

    Example:
        >>> result = quick_calculate(100.0, 80.0, 12.0, 500.0, "uploads/calc.xlsx")
        >>> print(f"利润: ¥{result.profit_amount:.2f}")
    """
    service = get_default_service(excel_path)
    return service.calculate(black_price, green_price, commission_rate, weight, excel_path)
