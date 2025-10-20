#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析Excel文件中的计算器逻辑
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import sys
import os

def analyze_excel_file(file_path):
    """分析Excel文件的结构和计算逻辑"""
    
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return
    
    print(f"正在分析Excel文件: {file_path}")
    print("=" * 60)
    
    try:
        # 使用openpyxl读取工作簿
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        print(f"工作表数量: {len(workbook.sheetnames)}")
        print(f"工作表名称: {workbook.sheetnames}")
        print()
        
        # 分析每个工作表
        for sheet_name in workbook.sheetnames:
            print(f"分析工作表: {sheet_name}")
            print("-" * 40)
            
            worksheet = workbook[sheet_name]
            
            # 获取工作表的使用范围
            if worksheet.max_row > 0 and worksheet.max_column > 0:
                print(f"数据范围: A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}")
                
                # 查找包含公式的单元格
                formulas = []
                for row in range(1, min(worksheet.max_row + 1, 100)):  # 限制在前100行
                    for col in range(1, min(worksheet.max_column + 1, 50)):  # 限制在前50列
                        cell = worksheet.cell(row=row, column=col)
                        if cell.value is not None:
                            if str(cell.value).startswith('='):
                                formulas.append({
                                    'cell': f"{get_column_letter(col)}{row}",
                                    'formula': cell.value,
                                    'value': cell.displayed_value if hasattr(cell, 'displayed_value') else 'N/A'
                                })
                
                if formulas:
                    print(f"\n发现 {len(formulas)} 个公式单元格:")
                    for formula in formulas[:20]:  # 只显示前20个公式
                        print(f"  {formula['cell']}: {formula['formula']}")
                
                # 显示前几行数据以了解结构
                print(f"\n前10行数据预览:")
                for row in range(1, min(11, worksheet.max_row + 1)):
                    row_data = []
                    for col in range(1, min(worksheet.max_column + 1, 10)):  # 只显示前10列
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value is not None:
                            row_data.append(str(cell_value)[:20])  # 限制显示长度
                        else:
                            row_data.append("")
                    if any(row_data):  # 只显示非空行
                        print(f"  第{row}行: {' | '.join(row_data)}")
            
            print()
        
        # 使用pandas读取数据进行进一步分析
        try:
            print("使用pandas读取数据:")
            print("-" * 40)
            
            # 读取所有工作表
            excel_data = pd.read_excel(file_path, sheet_name=None, header=None)
            
            for sheet_name, df in excel_data.items():
                print(f"\n工作表 '{sheet_name}' 数据形状: {df.shape}")
                
                # 显示非空数据的概览
                non_empty_df = df.dropna(how='all').dropna(axis=1, how='all')
                if not non_empty_df.empty:
                    print(f"有效数据范围: {non_empty_df.shape}")
                    print("前几行数据:")
                    print(non_empty_df.head(10).to_string())
                
        except Exception as e:
            print(f"pandas读取出错: {e}")
    
    except Exception as e:
        print(f"分析Excel文件时出错: {e}")

if __name__ == "__main__":
    file_path = "docs/122义乌仓资费测算3.17.xlsx"
    analyze_excel_file(file_path)